import base64
import csv
import io
import os
import random
import re
import uuid
from datetime import datetime, date, time, timezone, timedelta
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, status, Request, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload, selectinload
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pypdf

from app.core.database import get_db
from app.core.security import get_current_user, RoleChecker, get_password_hash
from app.core.utils import (
    calculate_haversine_distance, is_wfh_active, log_audit, get_safe_timezone,
    validate_image_bytes
)
from app.models.models import User, Attendance, BreakSession, OfficeSetting, AttendanceCorrectionRequest, EmployeeProfile
from app.schemas.schemas import (
    AttendanceOut, ClockInRequest, AttendanceCorrectionCreate, AttendanceCorrectionReview,
    AttendanceCorrectionOut, AttendanceImportResponse
)
from app.core.config import settings as app_settings

router = APIRouter(prefix="/attendance", tags=["Attendance Management"])

admin_required = RoleChecker(["Admin"])

@router.get("/today", response_model=Optional[AttendanceOut])
def get_today_attendance(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    today = date.today()
    attendance = db.query(Attendance).options(
        selectinload(Attendance.break_sessions)
    ).filter(
        Attendance.organization_id == current_user.organization_id,
        Attendance.user_id == current_user.id,
        Attendance.date == today
    ).first()
    return attendance

DEFAULT_HISTORY_LIMIT = 500
MAX_HISTORY_LIMIT = 2000

@router.get("/history", response_model=List[AttendanceOut])
def get_attendance_history(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    user_id: Optional[str] = None,
    status_filter: Optional[str] = None,
    limit: int = Query(DEFAULT_HISTORY_LIMIT, ge=1, le=MAX_HISTORY_LIMIT),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Attendance records, newest first.

    AttendanceOut nests break_sessions, so without the selectinload below
    serialising N records costs N+1 queries - an unfiltered admin call over
    8,000 records measured 8,015 queries and ~14s. The result set is also
    capped: this endpoint previously returned an organization's entire
    attendance history (5.8 MB at that size) on every call. Callers that need
    an older window should pass start_date/end_date or page with offset.
    """
    query = db.query(Attendance).options(
        selectinload(Attendance.break_sessions)
    ).filter(Attendance.organization_id == current_user.organization_id)

    if current_user.role == "Employee":
        query = query.filter(Attendance.user_id == current_user.id)
    else:
        if user_id:
            try:
                target_uuid = uuid.UUID(user_id) if isinstance(user_id, str) else user_id
                query = query.filter(Attendance.user_id == target_uuid)
            except (ValueError, TypeError):
                pass

    if start_date:
        query = query.filter(Attendance.date >= start_date)
    if end_date:
        query = query.filter(Attendance.date <= end_date)
    if status_filter:
        query = query.filter(Attendance.status == status_filter)

    return (
        query.order_by(Attendance.date.desc(), Attendance.clock_in.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )

@router.post("/clock-in", response_model=AttendanceOut)
def clock_in(
    request: Request,
    payload: ClockInRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    today = date.today()
    now_utc = datetime.now(timezone.utc)

    # Prevent duplicate clock-in
    existing = db.query(Attendance).filter(
        Attendance.organization_id == current_user.organization_id,
        Attendance.user_id == current_user.id,
        Attendance.date == today
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You have already clocked in for today"
        )

    # Fetch Tenant Office Settings
    settings = db.query(OfficeSetting).filter(OfficeSetting.organization_id == current_user.organization_id).first()
    if not settings:
        settings = OfficeSetting(organization_id=current_user.organization_id)
        db.add(settings)
        db.commit()
        db.refresh(settings)

    # Enforce Selfie check for Employees, unless this organization has turned
    # photo verification off in its office settings. Admins have always been
    # exempt. A selfie sent while the requirement is off is still accepted and
    # stored — turning the setting off relaxes the requirement, it does not
    # reject evidence an employee chose to provide.
    if settings.require_selfie and current_user.role == "Employee" and not payload.selfie_base64:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Selfie verification is required to clock in."
        )

    # Save Selfie image
    selfie_url = None
    if payload.selfie_base64:
        try:
            base64_data = payload.selfie_base64
            if "," in base64_data:
                _, base64_data = base64_data.split(",", 1)
            image_bytes = base64.b64decode(base64_data)

            if len(image_bytes) > app_settings.MAX_UPLOAD_SIZE_BYTES:
                raise HTTPException(status_code=400, detail="Selfie image exceeds maximum allowed size (5MB)")

            try:
                validate_image_bytes(image_bytes)
            except ValueError as e:
                raise HTTPException(status_code=400, detail=f"Invalid selfie image: {e}")

            selfies_dir = os.path.join(app_settings.UPLOAD_DIR, "selfies")
            os.makedirs(selfies_dir, exist_ok=True)
            
            filename = f"{current_user.id}_{today}_{uuid.uuid4().hex[:8]}.jpg"
            filepath = os.path.join(selfies_dir, filename)
            
            with open(filepath, "wb") as f:
                f.write(image_bytes)
            
            selfie_url = f"/api/static/selfies/{filename}"
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Failed to process selfie image verification: {str(e)}"
            )

    # Determine WFH status
    profile = current_user.profile
    wfh_active = is_wfh_active(profile, today)

    # Verify GPS location if WFH is not active
    if not wfh_active:
        distance = calculate_haversine_distance(
            payload.latitude,
            payload.longitude,
            settings.latitude,
            settings.longitude
        )
        if distance > settings.allowed_radius:
            log_audit(db, current_user.id, "CLOCK_IN_FAILED_OUTSIDE_RADIUS", request.client.host if request.client else None, f"Distance: {distance:.2f}m", organization_id=current_user.organization_id)
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"You are outside office location. Distance: {distance:.1f} meters."
            )

    # Suspicious check (e.g. low GPS accuracy > 200m)
    is_suspicious = bool(payload.gps_accuracy and payload.gps_accuracy > 200.0)

    # Calculate late minutes based on local time of the office timezone
    office_tz = get_safe_timezone(settings.timezone)
    local_now = datetime.now(office_tz)
    office_start_dt = datetime.combine(today, settings.office_start_time).replace(tzinfo=office_tz)
    
    late_minutes = 0
    if local_now > office_start_dt:
        late_diff = local_now - office_start_dt
        late_minutes = int(late_diff.total_seconds() / 60)

    # Set status
    if wfh_active:
        status_str = "Work From Home"
    elif late_minutes > 15:  # 15 minutes grace period
        status_str = "Late"
    else:
        status_str = "Present"

    new_attendance = Attendance(
        organization_id=current_user.organization_id,
        user_id=current_user.id,
        date=today,
        clock_in=now_utc,
        status=status_str,
        is_wfh=wfh_active,
        late_minutes=late_minutes,
        selfie_url=selfie_url,
        latitude=payload.latitude,
        longitude=payload.longitude,
        gps_accuracy=payload.gps_accuracy,
        device_info=payload.device_info,
        is_suspicious=is_suspicious
    )
    
    db.add(new_attendance)
    db.commit()
    db.refresh(new_attendance)

    log_audit(db, current_user.id, "CLOCK_IN_SUCCESS", request.client.host if request.client else None, f"Status: {status_str}, Late: {late_minutes} min", organization_id=current_user.organization_id)
    return new_attendance

@router.post("/clock-out", response_model=AttendanceOut)
def clock_out(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    today = date.today()
    now_utc = datetime.now(timezone.utc)

    attendance = db.query(Attendance).filter(
        Attendance.organization_id == current_user.organization_id,
        Attendance.user_id == current_user.id,
        Attendance.date == today
    ).first()

    if not attendance:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You must clock in first before clocking out"
        )

    if attendance.clock_out:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You have already clocked out for today"
        )

    # Close any open break session
    open_break = db.query(BreakSession).filter(
        BreakSession.attendance_id == attendance.id,
        BreakSession.end_time == None
    ).first()
    if open_break:
        open_break.end_time = now_utc
        start_tz = open_break.start_time if open_break.start_time.tzinfo else open_break.start_time.replace(tzinfo=timezone.utc)
        open_break.duration = (now_utc - start_tz).total_seconds() / 60.0

    attendance.clock_out = now_utc

    settings = db.query(OfficeSetting).filter(OfficeSetting.organization_id == current_user.organization_id).first()
    if not settings:
        # Must be persisted, not left transient: Column(default=...) only
        # applies at INSERT, so an un-flushed OfficeSetting() has None for
        # every configured field and the calculations below crash on it.
        # Every other call site already commits here.
        settings = OfficeSetting(organization_id=current_user.organization_id)
        db.add(settings)
        db.commit()
        db.refresh(settings)

    # Calculate total duration in hours
    clock_in_tz = attendance.clock_in if attendance.clock_in.tzinfo else attendance.clock_in.replace(tzinfo=timezone.utc)
    total_seconds = (now_utc - clock_in_tz).total_seconds()
    total_hours = max(0.0, total_seconds / 3600.0)

    # Calculate breaks. flush (not commit) is enough to make the break closed
    # above visible to this aggregate, and keeps the whole clock-out in one
    # transaction instead of committing a half-finished record partway through.
    db.flush()
    total_break_minutes = db.query(
        func.coalesce(func.sum(BreakSession.duration), 0.0)
    ).filter(BreakSession.attendance_id == attendance.id).scalar()
    attendance.break_duration = total_break_minutes / 60.0

    # Net working hours
    attendance.working_hours = max(0.0, total_hours - attendance.break_duration)

    # Early leaving calculation
    office_tz = get_safe_timezone(settings.timezone)
    local_now = datetime.now(office_tz)
    office_end_dt = datetime.combine(today, settings.office_end_time).replace(tzinfo=office_tz)
    early_leaving_minutes = 0
    if local_now < office_end_dt:
        early_diff = office_end_dt - local_now
        early_leaving_minutes = int(early_diff.total_seconds() / 60)
    attendance.early_leaving_minutes = early_leaving_minutes

    # Overtime calculation
    overtime_minutes = 0
    net_working_minutes = attendance.working_hours * 60.0
    required_minutes = settings.required_working_hours * 60.0
    if net_working_minutes > required_minutes:
        overtime_minutes = int(net_working_minutes - required_minutes)
    attendance.overtime_minutes = overtime_minutes

    # Recalculate status if working hours indicate a Half Day
    if attendance.working_hours < (settings.required_working_hours / 2.0):
        attendance.status = "Half Day"

    db.commit()
    db.refresh(attendance)

    log_audit(db, current_user.id, "CLOCK_OUT_SUCCESS", request.client.host if request.client else None, f"Hours: {attendance.working_hours:.2f}, OT: {overtime_minutes} min", organization_id=current_user.organization_id)
    return attendance

@router.post("/break/start")
def start_break(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    today = date.today()
    now_utc = datetime.now(timezone.utc)

    attendance = db.query(Attendance).filter(
        Attendance.organization_id == current_user.organization_id,
        Attendance.user_id == current_user.id,
        Attendance.date == today
    ).first()

    if not attendance:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You must clock in first before starting a break"
        )
        
    if attendance.clock_out:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot start break after clocking out"
        )

    active_break = db.query(BreakSession).filter(
        BreakSession.attendance_id == attendance.id,
        BreakSession.end_time == None
    ).first()
    
    if active_break:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You are already on a break"
        )

    new_break = BreakSession(
        attendance_id=attendance.id,
        start_time=now_utc
    )
    db.add(new_break)
    db.commit()
    
    log_audit(db, current_user.id, "BREAK_START", request.client.host if request.client else None, organization_id=current_user.organization_id)
    return {"message": "Break started successfully", "start_time": now_utc}

@router.post("/break/end")
def end_break(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    today = date.today()
    now_utc = datetime.now(timezone.utc)

    attendance = db.query(Attendance).filter(
        Attendance.organization_id == current_user.organization_id,
        Attendance.user_id == current_user.id,
        Attendance.date == today
    ).first()

    if not attendance:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Attendance record not found for today"
        )

    active_break = db.query(BreakSession).filter(
        BreakSession.attendance_id == attendance.id,
        BreakSession.end_time == None
    ).first()

    if not active_break:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You are not on a break currently"
        )

    start_tz = active_break.start_time if active_break.start_time.tzinfo else active_break.start_time.replace(tzinfo=timezone.utc)
    active_break.end_time = now_utc
    active_break.duration = (now_utc - start_tz).total_seconds() / 60.0
    db.flush()

    # Recalculate break duration on main attendance. Summed in the database
    # rather than by loading every break row for the day.
    total_break_minutes = db.query(
        func.coalesce(func.sum(BreakSession.duration), 0.0)
    ).filter(BreakSession.attendance_id == attendance.id).scalar()
    attendance.break_duration = total_break_minutes / 60.0
    db.commit()

    log_audit(db, current_user.id, "BREAK_END", request.client.host if request.client else None, f"Duration: {active_break.duration:.1f} min", organization_id=current_user.organization_id)
    return {"message": "Break ended successfully", "duration_minutes": active_break.duration}

@router.post("/corrections", response_model=AttendanceCorrectionOut)
def create_correction_request(
    payload: AttendanceCorrectionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    existing = db.query(AttendanceCorrectionRequest).filter(
        AttendanceCorrectionRequest.organization_id == current_user.organization_id,
        AttendanceCorrectionRequest.user_id == current_user.id,
        AttendanceCorrectionRequest.date == payload.date,
        AttendanceCorrectionRequest.status == "Pending"
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"You already have a pending correction request for {payload.date}."
        )
    
    new_request = AttendanceCorrectionRequest(
        organization_id=current_user.organization_id,
        user_id=current_user.id,
        date=payload.date,
        proposed_clock_in=payload.proposed_clock_in,
        proposed_clock_out=payload.proposed_clock_out,
        reason=payload.reason,
        status="Pending"
    )
    db.add(new_request)
    db.commit()
    db.refresh(new_request)
    return new_request

@router.get("/corrections", response_model=List[AttendanceCorrectionOut])
def get_correction_requests(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # AttendanceCorrectionOut nests user -> profile -> department; eager load
    # them so serialising the list stays a fixed number of queries.
    query = db.query(AttendanceCorrectionRequest).options(
        joinedload(AttendanceCorrectionRequest.user)
        .joinedload(User.profile)
        .joinedload(EmployeeProfile.department)
    ).filter(AttendanceCorrectionRequest.organization_id == current_user.organization_id)
    if current_user.role == "Employee":
        query = query.filter(AttendanceCorrectionRequest.user_id == current_user.id)
    return query.order_by(AttendanceCorrectionRequest.created_at.desc()).all()

@router.patch("/corrections/{id}/review", response_model=AttendanceCorrectionOut)
def review_correction_request(
    id: str,
    payload: AttendanceCorrectionReview,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    _admin_only = Depends(admin_required)
):
    import uuid as py_uuid
    try:
        corr_uuid = py_uuid.UUID(id) if isinstance(id, str) else id
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid correction request ID format")

    corr = db.query(AttendanceCorrectionRequest).filter(
        AttendanceCorrectionRequest.id == corr_uuid,
        AttendanceCorrectionRequest.organization_id == current_user.organization_id
    ).first()
    if not corr:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Correction request not found."
        )
    
    if corr.status != "Pending":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Request is already {corr.status}."
        )
    
    corr.status = payload.status
    corr.comment = payload.comment
    
    if payload.status == "Approved":
        attendance = db.query(Attendance).filter(
            Attendance.organization_id == current_user.organization_id,
            Attendance.user_id == corr.user_id,
            Attendance.date == corr.date
        ).first()
        
        settings = db.query(OfficeSetting).filter(OfficeSetting.organization_id == current_user.organization_id).first()
        if not settings:
            settings = OfficeSetting(organization_id=current_user.organization_id)
            db.add(settings)
            db.commit()
            db.refresh(settings)
            
        wfh_active = is_wfh_active(corr.user.profile if corr.user else None, corr.date)
        
        c_in = corr.proposed_clock_in if corr.proposed_clock_in else (attendance.clock_in if attendance else datetime.combine(corr.date, settings.office_start_time).replace(tzinfo=timezone.utc))
        c_out = corr.proposed_clock_out if corr.proposed_clock_out else (attendance.clock_out if attendance else datetime.combine(corr.date, settings.office_end_time).replace(tzinfo=timezone.utc))
        
        total_hours = 0.0
        if c_in and c_out:
            c_in_t = c_in if c_in.tzinfo else c_in.replace(tzinfo=timezone.utc)
            c_out_t = c_out if c_out.tzinfo else c_out.replace(tzinfo=timezone.utc)
            total_hours = max(0.0, (c_out_t - c_in_t).total_seconds() / 3600.0)
            
        break_dur = attendance.break_duration if attendance else settings.lunch_break_hours
        working_hours = max(0.0, total_hours - break_dur)
        
        office_tz = get_safe_timezone(settings.timezone)
        
        c_in_local = c_in.astimezone(office_tz) if c_in.tzinfo else c_in.replace(tzinfo=timezone.utc).astimezone(office_tz)
        office_start_dt = datetime.combine(corr.date, settings.office_start_time).replace(tzinfo=office_tz)
        
        late_minutes = 0
        if c_in_local > office_start_dt:
            late_diff = c_in_local - office_start_dt
            late_minutes = int(late_diff.total_seconds() / 60)
            
        c_out_local = c_out.astimezone(office_tz) if c_out.tzinfo else c_out.replace(tzinfo=timezone.utc).astimezone(office_tz)
        office_end_dt = datetime.combine(corr.date, settings.office_end_time).replace(tzinfo=office_tz)
        early_leaving_minutes = 0
        if c_out_local < office_end_dt:
            early_diff = office_end_dt - c_out_local
            early_leaving_minutes = int(early_diff.total_seconds() / 60)
            
        overtime_minutes = 0
        required_minutes = settings.required_working_hours * 60.0
        if (working_hours * 60.0) > required_minutes:
            overtime_minutes = int((working_hours * 60.0) - required_minutes)
            
        if wfh_active:
            status_str = "Work From Home"
        elif working_hours < (settings.required_working_hours / 2.0):
            status_str = "Half Day"
        elif late_minutes > 15:
            status_str = "Late"
        else:
            status_str = "Present"
            
        if not attendance:
            attendance = Attendance(
                organization_id=current_user.organization_id,
                user_id=corr.user_id,
                date=corr.date,
                clock_in=c_in,
                clock_out=c_out,
                working_hours=working_hours,
                break_duration=break_dur,
                status=status_str,
                is_wfh=wfh_active,
                late_minutes=late_minutes,
                early_leaving_minutes=early_leaving_minutes,
                overtime_minutes=overtime_minutes,
                modified_by_admin=True,
                modification_reason=corr.reason
            )
            db.add(attendance)
        else:
            attendance.clock_in = c_in
            attendance.clock_out = c_out
            attendance.working_hours = working_hours
            attendance.break_duration = break_dur
            attendance.status = status_str
            attendance.is_wfh = wfh_active
            attendance.late_minutes = late_minutes
            attendance.early_leaving_minutes = early_leaving_minutes
            attendance.overtime_minutes = overtime_minutes
            attendance.modified_by_admin = True
            attendance.modification_reason = corr.reason
            
    db.commit()
    db.refresh(corr)
    
    log_action = "CORRECTION_APPROVED" if payload.status == "Approved" else "CORRECTION_REJECTED"
    log_audit(db, current_user.id, log_action, request.client.host if request.client else None, f"Request ID: {corr.id}, Employee ID: {corr.user_id}", organization_id=current_user.organization_id)
    return corr

# --- Attendance Bulk Import & Template Generation ---

def parse_date_value(val: Any) -> Optional[date]:
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val_str = val.strip().split("\n")[0].strip()
        m = re.search(r"(\d{1,4})[./-](\d{1,2})[./-](\d{1,4})", val_str)
        if m:
            p1, p2, p3 = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if p1 > 1900:
                try:
                    return date(p1, p2, p3)
                except ValueError:
                    pass
            elif p3 > 1900:
                try:
                    return date(p3, p2, p1)
                except ValueError:
                    try:
                        return date(p3, p1, p2)
                    except ValueError:
                        pass
        for fmt in (
            "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y", "%Y%m%d",
            "%d %B %Y", "%d %b %Y", "%B %d, %Y", "%b %d, %Y", "%d %B", "%d %b"
        ):
            try:
                return datetime.strptime(val_str, fmt).date()
            except ValueError:
                continue
    return None

def parse_time_value(val: Any, target_date: date, is_clock_out: bool = False, ref_clock_in: Optional[datetime] = None) -> Optional[datetime]:
    if not val:
        return None
    if isinstance(val, datetime):
        h, m, s = val.hour, val.minute, val.second
        if is_clock_out and 1 <= h <= 11:
            h += 12
        return datetime(target_date.year, target_date.month, target_date.day, h, m, s, tzinfo=timezone.utc)
    if isinstance(val, time):
        h, m, s = val.hour, val.minute, val.second
        if is_clock_out and 1 <= h <= 11:
            h += 12
        return datetime(target_date.year, target_date.month, target_date.day, h, m, s, tzinfo=timezone.utc)
    if isinstance(val, (int, float)):
        h = int(val)
        m = int(round((val - h) * 60))
        if h < 0 or h > 23 or m < 0 or m > 59:
            return None
        if is_clock_out and 1 <= h <= 11:
            h += 12
        return datetime(target_date.year, target_date.month, target_date.day, h, m, 0, tzinfo=timezone.utc)
    if isinstance(val, str):
        val_str = val.strip()
        m = re.search(r"(\d{1,2})[:.](\d{2})(?:[:.](\d{2}))?\s*(am|pm)?", val_str, re.IGNORECASE)
        if m:
            h = int(m.group(1))
            min_val = int(m.group(2))
            sec_val = int(m.group(3)) if m.group(3) else 0
            ampm = m.group(4).lower() if m.group(4) else None

            if ampm:
                if h < 1 or h > 12 or min_val < 0 or min_val > 59 or sec_val < 0 or sec_val > 59:
                    return None
                if ampm == "pm" and h < 12:
                    h += 12
                elif ampm == "am" and h == 12:
                    h = 0
            else:
                if h < 0 or h > 23 or min_val < 0 or min_val > 59 or sec_val < 0 or sec_val > 59:
                    return None
                if is_clock_out and 1 <= h <= 11:
                    h += 12

            try:
                return datetime(target_date.year, target_date.month, target_date.day, h, min_val, sec_val, tzinfo=timezone.utc)
            except ValueError:
                return None

        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p", "%I:%M%p", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                dt = datetime.strptime(val_str, fmt)
                h = dt.hour
                if is_clock_out and 1 <= h <= 11 and "p" not in val_str.lower() and "a" not in val_str.lower():
                    h += 12
                return datetime(target_date.year, target_date.month, target_date.day, h, dt.minute, dt.second, tzinfo=timezone.utc)
            except ValueError:
                continue
    return None

def parse_float_value(val: Any, default: float = 0.0) -> float:
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def _is_tracker_header(row_vals: List[Any]) -> bool:
    row_text = " ".join(str(v or "") for v in row_vals).lower()
    has_tracker_words = any(w in row_text for w in ["sign in", "current date", "lunch time", "task responsibility", "today task", "yesterday task", "punch in"])
    has_emp = any(w in row_text for w in ["employee", "responsibility", "name"])
    return has_emp and has_tracker_words

def _try_parse_daily_tracker_rows(raw_rows: List[List[Any]]) -> Optional[List[Dict[str, Any]]]:
    header_indices = [i for i, r in enumerate(raw_rows) if _is_tracker_header(r)]
    if not header_indices:
        return None

    parsed_records: List[Dict[str, Any]] = []
    current_date: Optional[date] = None
    col_map = {"emp": 1, "date": 2, "cin": 3, "cout": 4, "lout": 5, "lin": 6, "status": 9}

    for row_idx, row in enumerate(raw_rows, start=1):
        if not row or not any(c is not None and str(c).strip() != "" for c in row):
            continue

        if _is_tracker_header(row):
            for idx, c in enumerate(row):
                c_str = str(c or "").strip().lower()
                if "employee" in c_str or "responsibility" in c_str or c_str == "name":
                    col_map["emp"] = idx
                elif "date" in c_str or "day" in c_str:
                    col_map["date"] = idx
                elif any(k in c_str for k in ["sign in", "in time", "clock in", "punch in"]):
                    col_map["cin"] = idx
                elif any(k in c_str for k in ["sign out", "out time", "clock out", "punch out"]):
                    col_map["cout"] = idx
                elif "lunch" in c_str and "out" in c_str:
                    col_map["lout"] = idx
                elif "lunch" in c_str and "in" in c_str:
                    col_map["lin"] = idx
                elif "status" in c_str or "stetus" in c_str or "task" in c_str:
                    col_map["status"] = idx
            continue

        for cell in row[:5]:
            d = parse_date_value(cell)
            if d:
                current_date = d
                break

        emp_idx = col_map.get("emp", 1)
        if emp_idx < len(row):
            emp_raw = row[emp_idx]
            emp_str = str(emp_raw or "").strip()
            if emp_str and not emp_str.isdigit() and emp_str.lower() not in ["#", "employee name", "name", "task responsibility", "total", "task"]:
                cin_raw = row[col_map["cin"]] if col_map.get("cin") is not None and col_map["cin"] < len(row) else None
                cout_raw = row[col_map["cout"]] if col_map.get("cout") is not None and col_map["cout"] < len(row) else None
                lout_raw = row[col_map["lout"]] if col_map.get("lout") is not None and col_map["lout"] < len(row) else None
                lin_raw = row[col_map["lin"]] if col_map.get("lin") is not None and col_map["lin"] < len(row) else None
                status_raw = row[col_map["status"]] if col_map.get("status") is not None and col_map["status"] < len(row) else None

                if current_date:
                    parsed_records.append({
                        "row_num": row_idx,
                        "employee_raw": emp_str,
                        "date_raw": current_date,
                        "clock_in_raw": cin_raw,
                        "clock_out_raw": cout_raw,
                        "lunch_out_raw": lout_raw,
                        "lunch_in_raw": lin_raw,
                        "hours_raw": None,
                        "status_raw": str(status_raw).strip() if status_raw is not None else "",
                        "overtime_raw": 0
                    })

    return parsed_records if len(parsed_records) > 0 else None

def parse_attendance_file_rows(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    ext = os.path.splitext(filename)[1].lower()
    raw_rows: List[List[Any]] = []

    if ext in [".xlsx", ".xls"]:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            # Find candidate sheet
            target_sheet = None
            for sname in wb.sheetnames:
                s_lower = sname.lower()
                if any(k in s_lower for k in ["daily tracker new", "daily tracker", "attendance", "timesheet"]):
                    target_sheet = wb[sname]
                    break
            if target_sheet is None:
                target_sheet = wb.active

            for row in target_sheet.iter_rows(values_only=True):
                if any(c is not None and str(c).strip() != "" for c in row):
                    raw_rows.append(list(row))

            tracker_records = _try_parse_daily_tracker_rows(raw_rows)
            if tracker_records:
                return tracker_records
        except Exception as e:
            if isinstance(e, ValueError) and "no data rows" in str(e):
                raise
            raise ValueError(f"Failed to read Excel workbook: {str(e)}")

    elif ext in [".csv", ".txt"]:
        try:
            text = file_bytes.decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                if any(c.strip() for c in row):
                    raw_rows.append(row)
            tracker_records = _try_parse_daily_tracker_rows(raw_rows)
            if tracker_records:
                return tracker_records
        except Exception as e:
            if isinstance(e, ValueError) and "no data rows" in str(e):
                raise
            raise ValueError(f"Failed to read CSV text: {str(e)}")

    elif ext == ".pdf":
        try:
            reader = pypdf.PdfReader(io.BytesIO(file_bytes))
            for page in reader.pages:
                txt = page.extract_text()
                if txt:
                    for line in txt.split("\n"):
                        parts = [p.strip() for p in re.split(r"[,|\t]+|\s{2,}", line.strip()) if p.strip()]
                        if len(parts) >= 2:
                            raw_rows.append(parts)
        except Exception as e:
            raise ValueError(f"Failed to parse PDF attendance document: {str(e)}")

    else:
        raise ValueError(f"Unsupported file format '{ext}'. Allowed formats: .xlsx, .xls, .csv, .pdf, .txt")

    if len(raw_rows) < 2:
        raise ValueError("Attendance file is empty or contains no data rows.")

    # Map header column names to field keys
    field_map: Dict[str, int] = {}
    for idx, col in enumerate(raw_rows[0]):
        c_str = str(col).strip().lower() if col is not None else ""
        if any(k in c_str for k in ["employee_id", "emp_id", "empid", "employee id", "emp id", "code", "emp_code", "employee code", "email", "user"]):
            field_map["employee"] = idx
        elif any(k in c_str for k in ["date", "day", "attendance_date"]):
            field_map["date"] = idx
        elif any(k in c_str for k in ["clock_in", "clock in", "punch in", "in_time", "in time", "start"]):
            field_map["clock_in"] = idx
        elif any(k in c_str for k in ["clock_out", "clock out", "punch out", "out_time", "out time", "end"]):
            field_map["clock_out"] = idx
        elif any(k in c_str for k in ["working_hours", "working hours", "work hours", "net hours", "hours", "total hours"]):
            field_map["working_hours"] = idx
        elif any(k in c_str for k in ["status", "state", "attendance_status"]):
            field_map["status"] = idx
        elif any(k in c_str for k in ["overtime", "ot", "overtime_minutes"]):
            field_map["overtime"] = idx

    # Fallback to positional mapping if headers are non-standard
    if "employee" not in field_map:
        field_map["employee"] = 0
    if "date" not in field_map and len(raw_rows[0]) > 1:
        field_map["date"] = 1

    parsed_records = []
    for row_num, row in enumerate(raw_rows[1:], start=2):
        if not any(c is not None and str(c).strip() != "" for c in row):  # pragma: no cover
            continue

        emp_raw = row[field_map["employee"]] if field_map.get("employee") is not None and field_map["employee"] < len(row) else None
        date_raw = row[field_map["date"]] if field_map.get("date") is not None and field_map["date"] < len(row) else None
        cin_raw = row[field_map["clock_in"]] if field_map.get("clock_in") is not None and field_map["clock_in"] < len(row) else None
        cout_raw = row[field_map["clock_out"]] if field_map.get("clock_out") is not None and field_map["clock_out"] < len(row) else None
        hours_raw = row[field_map["working_hours"]] if field_map.get("working_hours") is not None and field_map["working_hours"] < len(row) else None
        status_raw = row[field_map["status"]] if field_map.get("status") is not None and field_map["status"] < len(row) else None
        ot_raw = row[field_map["overtime"]] if field_map.get("overtime") is not None and field_map["overtime"] < len(row) else None

        parsed_records.append({
            "row_num": row_num,
            "employee_raw": str(emp_raw).strip() if emp_raw is not None else "",
            "date_raw": date_raw,
            "clock_in_raw": cin_raw,
            "clock_out_raw": cout_raw,
            "hours_raw": hours_raw,
            "status_raw": str(status_raw).strip() if status_raw is not None else "",
            "overtime_raw": ot_raw
        })

    return parsed_records

@router.get("/template")
def download_attendance_import_template(
    admin_user: User = Depends(admin_required)
):
    """
    Generates a pre-formatted Excel workbook template for bulk employee attendance imports.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Import Template"

    header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    headers = [
        "Employee ID / Email",
        "Date (YYYY-MM-DD)",
        "Clock In (HH:MM)",
        "Clock Out (HH:MM)",
        "Working Hours",
        "Status",
        "Overtime (Mins)"
    ]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    sample_rows = [
        ["EMP-1001", "2026-08-01", "09:30", "18:30", 8.0, "Present", 0],
        ["EMP-1002", "2026-08-01", "09:45", "18:30", 7.75, "Late", 0],
        ["EMP-1003", "2026-08-01", "09:30", "18:00", 8.0, "Work From Home", 0],
        ["EMP-1004", "2026-08-01", "", "", 0.0, "Leave", 0],
        ["EMP-1001", "2026-08-02", "09:30", "20:00", 9.5, "Present", 60],
    ]

    for row_data in sample_rows:
        ws.append(row_data)

    for row_idx in range(2, len(sample_rows) + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = left_align if col_idx == 1 else center_align

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 16)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "attendance_import_template.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/import", response_model=AttendanceImportResponse)
def import_attendance_file(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin_user: User = Depends(admin_required)
):
    """
    Parses and bulk-imports attendance records from Excel, CSV, or PDF files.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file selected for import")

    file_bytes = file.file.read()
    if len(file_bytes) == 0:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if len(file_bytes) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File exceeds maximum size of 10 MB")

    try:
        raw_records = parse_attendance_file_rows(file_bytes, file.filename)
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))

    # Pre-fetch all active organization employees for fast in-memory lookup
    org_users = db.query(User).options(joinedload(User.profile)).filter(
        User.organization_id == admin_user.organization_id
    ).all()

    file_dates = {d for d in (parse_date_value(r["date_raw"]) for r in raw_records) if d}
    existing_by_key: Dict[Any, Attendance] = {}
    if file_dates:
        existing_by_key = {
            (att.user_id, att.date): att
            for att in db.query(Attendance).filter(
                Attendance.organization_id == admin_user.organization_id,
                Attendance.date >= min(file_dates),
                Attendance.date <= max(file_dates),
            )
        }

    emp_lookup: Dict[str, User] = {}
    for u in org_users:
        if u.email:
            emp_lookup[u.email.lower().strip()] = u
        if u.profile:
            if u.profile.employee_id:
                emp_lookup[u.profile.employee_id.lower().strip()] = u
                emp_lookup[u.profile.employee_id.upper().strip()] = u
            if u.profile.first_name:
                first = u.profile.first_name.lower().strip()
                emp_lookup[first] = u
                if u.profile.last_name:
                    last = u.profile.last_name.lower().strip()
                    emp_lookup[f"{first} {last}"] = u

    imported_count = 0
    updated_count = 0
    skipped_count = 0
    errors: List[str] = []

    for row in raw_records:
        r_num = row["row_num"]
        emp_key = row["employee_raw"].strip()
        if not emp_key:
            skipped_count += 1
            errors.append(f"Row {r_num}: Missing employee ID or email")
            continue

        target_user = emp_lookup.get(emp_key.lower()) or emp_lookup.get(emp_key.upper())
        if not target_user:
            skipped_count += 1
            errors.append(f"Row {r_num}: Employee '{emp_key}' not found in your organization")
            continue

        rec_date = parse_date_value(row["date_raw"])
        if not rec_date:
            skipped_count += 1
            errors.append(f"Row {r_num}: Invalid or missing date '{row['date_raw']}'")
            continue

        # Parse clock in / out
        c_in = parse_time_value(row["clock_in_raw"], rec_date, is_clock_out=False)
        c_out = parse_time_value(row["clock_out_raw"], rec_date, is_clock_out=True, ref_clock_in=c_in)

        # Lunch break duration calculation
        break_mins = 0
        if row.get("lunch_out_raw") and row.get("lunch_in_raw"):
            l_out = parse_time_value(row["lunch_out_raw"], rec_date, is_clock_out=False)
            l_in = parse_time_value(row["lunch_in_raw"], rec_date, is_clock_out=False)
            if l_out and l_in:
                diff_mins = (l_in - l_out).total_seconds() / 60.0
                if 0 < diff_mins <= 300:
                    break_mins = int(diff_mins)

        # Working hours
        w_hours = parse_float_value(row["hours_raw"], 0.0)
        if w_hours <= 0.0 and c_in and c_out:
            diff_secs = (c_out - c_in).total_seconds()
            if break_mins > 0:
                diff_secs = max(0, diff_secs - (break_mins * 60))
            if diff_secs > 0:
                w_hours = round(diff_secs / 3600.0, 2)

        # Overtime
        ot_mins = int(parse_float_value(row["overtime_raw"], 0.0))

        # Status normalization
        combined_text = f"{row.get('clock_in_raw', '')} {row.get('clock_out_raw', '')} {row.get('status_raw', '')}".lower()
        status_input = row["status_raw"].strip().title() if row.get("status_raw") else ""
        if not status_input:
            if w_hours >= 7.5:
                status_input = "Present"
            elif 4.0 <= w_hours < 7.5:
                status_input = "Half Day"
            elif 0.0 < w_hours < 4.0:
                status_input = "Late"
            else:
                status_input = "Absent"

        if "Home" in status_input or "Wfh" in status_input or "wfh" in combined_text:
            status_str = "Work From Home"
            is_wfh = True
        elif "Leave" in status_input or "On Leave" in status_input:
            status_str = "Leave"
            is_wfh = False
        elif "Half" in status_input or "half-day" in combined_text:
            status_str = "Half Day"
            is_wfh = False
        elif "Late" in status_input:
            status_str = "Late"
            is_wfh = False
        elif "Absent" in status_input:
            status_str = "Absent"
            is_wfh = False
        else:
            status_str = "Present"
            is_wfh = False

        if not c_in and status_str in ["Present", "Work From Home", "Late", "Half Day"]:
            c_in = datetime.combine(rec_date, time(9, 30), tzinfo=timezone.utc)
        elif not c_in and status_str in ["Leave", "Absent"]:
            c_in = datetime.combine(rec_date, time(0, 0), tzinfo=timezone.utc)

        if not c_out and status_str in ["Present", "Work From Home", "Late", "Half Day"] and w_hours > 0:
            c_out = c_in + timedelta(hours=w_hours)

        # Check existing attendance record
        existing = existing_by_key.get((target_user.id, rec_date))

        if existing:
            existing.clock_in = c_in or existing.clock_in
            existing.clock_out = c_out
            existing.working_hours = w_hours
            existing.break_duration = break_mins or existing.break_duration
            existing.status = status_str
            existing.is_wfh = is_wfh
            existing.overtime_minutes = ot_mins
            existing.modified_by_admin = True
            existing.modification_reason = f"Imported from {file.filename}"
            updated_count += 1
        else:
            new_att = Attendance(
                organization_id=admin_user.organization_id,
                user_id=target_user.id,
                date=rec_date,
                clock_in=c_in or datetime.combine(rec_date, time(9, 30), tzinfo=timezone.utc),
                clock_out=c_out,
                working_hours=w_hours,
                break_duration=break_mins,
                status=status_str,
                is_wfh=is_wfh,
                overtime_minutes=ot_mins,
                modified_by_admin=True,
                modification_reason=f"Imported from {file.filename}"
            )
            db.add(new_att)
            existing_by_key[(target_user.id, rec_date)] = new_att
            imported_count += 1

    db.commit()

    log_audit(
        db,
        admin_user.id,
        "ATTENDANCE_IMPORTED",
        request.client.host if request.client else None,
        f"File: {file.filename}, Imported: {imported_count}, Updated: {updated_count}, Skipped: {skipped_count}",
        organization_id=admin_user.organization_id
    )

    return AttendanceImportResponse(
        total_rows=len(raw_records),
        imported_count=imported_count,
        updated_count=updated_count,
        skipped_count=skipped_count,
        errors=errors[:15]
    )

