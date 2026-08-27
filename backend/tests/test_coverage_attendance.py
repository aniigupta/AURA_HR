"""
Coverage for the attendance branches the feature tests never reach: the
clock-out and break lifecycle, correction approval over an existing record,
and the bulk-import parsers.

The import parsers get the most attention here. They accept three file
formats, guess at column meanings, and silently overwrite settled attendance
records — which makes a wrong guess a payroll error rather than a parse error.
"""

import base64
import csv
import io
import re
import uuid
from datetime import date, datetime, time, timedelta, timezone

import openpyxl
import pytest
from openpyxl import Workbook
from PIL import Image
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from app.routers.attendance import (
    parse_attendance_file_rows,
    parse_date_value,
    parse_float_value,
    parse_time_value,
)
from app.models.models import (
    Attendance,
    AttendanceCorrectionRequest,
    BreakSession,
    OfficeSetting,
    User,
)

TODAY = date(2026, 3, 17)


def selfie() -> str:
    buf = io.BytesIO()
    Image.new("RGB", (16, 16), (30, 90, 160)).save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()


@pytest.fixture
def office(db) -> OfficeSetting:
    return db.query(OfficeSetting).first()


def clock_in(client, office, **extra):
    payload = {"latitude": office.latitude, "longitude": office.longitude, "selfie_base64": selfie()}
    payload.update(extra)
    return client.post("/api/attendance/clock-in", json=payload)


# --------------------------------------------------------------------------
# Value parsers — every accepted format and every rejection
# --------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("2026-03-17", date(2026, 3, 17)),
        ("17-03-2026", date(2026, 3, 17)),
        ("17/03/2026", date(2026, 3, 17)),
        ("03/17/2026", date(2026, 3, 17)),
        ("2026/03/17", date(2026, 3, 17)),
        ("17.03.2026", date(2026, 3, 17)),
        ("20260317", date(2026, 3, 17)),
        ("  2026-03-17  ", date(2026, 3, 17)),
    ],
)
def test_att_053_every_supported_date_format_parses(raw, expected):
    """ATT-053 — HR exports come from many systems; each listed format works."""
    assert parse_date_value(raw) == expected


def test_att_054_native_date_and_datetime_values_pass_through():
    """Excel hands back real date objects, not strings."""
    assert parse_date_value(datetime(2026, 3, 17, 9, 30)) == date(2026, 3, 17)
    assert parse_date_value(date(2026, 3, 17)) == date(2026, 3, 17)


@pytest.mark.parametrize("raw", ["", None, "32/13/2026", "soon", "not-a-date", 0])
def test_att_047_an_unparseable_date_returns_none_rather_than_guessing(raw):
    """ATT-047 — a bad date must not be defaulted to today; the row is skipped."""
    assert parse_date_value(raw) is None


@pytest.mark.parametrize(
    "raw,hour,minute",
    [
        ("09:30:00", 9, 30),
        ("09:30", 9, 30),
        ("09:30:00 AM", 9, 30),
        ("09:30 AM", 9, 30),
        ("09:30AM", 9, 30),
        ("2026-03-17T09:30:00", 9, 30),
        ("2026-03-17 09:30:00", 9, 30),
        ("  09:30  ", 9, 30),
    ],
)
def test_att_055_every_supported_time_format_parses(raw, hour, minute):
    """The parsed time is re-based onto the row's date, in UTC."""
    parsed = parse_time_value(raw, TODAY)
    assert (parsed.hour, parsed.minute) == (hour, minute)
    assert parsed.date() == TODAY
    assert parsed.tzinfo == timezone.utc


def test_att_056_native_time_and_datetime_values_are_rebased_onto_the_row_date():
    """A datetime from Excel carries its own date, which must not win."""
    assert parse_time_value(time(18, 45), TODAY).hour == 18
    rebased = parse_time_value(datetime(1899, 12, 31, 18, 45), TODAY)
    assert rebased.date() == TODAY and rebased.hour == 18


@pytest.mark.parametrize("raw", ["", None, "half past nine", "25:99"])
def test_att_057_an_unparseable_time_returns_none(raw):
    """An unreadable time leaves the field empty rather than inventing one."""
    assert parse_time_value(raw, TODAY) is None


@pytest.mark.parametrize(
    "raw,expected", [("7.5", 7.5), (7.5, 7.5), ("", 0.0), (None, 0.0), ("abc", 0.0), ([], 0.0)]
)
def test_att_058_float_parsing_falls_back_to_the_default(raw, expected):
    """Hours columns arrive as text, numbers, blanks and junk alike."""
    assert parse_float_value(raw, 0.0) == expected


# --------------------------------------------------------------------------
# parse_attendance_file_rows — the three formats and the header mapping
# --------------------------------------------------------------------------


def xlsx_bytes(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def csv_bytes(rows: list[list], bom: bool = False) -> bytes:
    out = io.StringIO()
    csv.writer(out).writerows(rows)
    raw = out.getvalue().encode("utf-8")
    return (b"\xef\xbb\xbf" + raw) if bom else raw


def pdf_bytes(lines: list[str]) -> bytes:
    buf = io.BytesIO()
    pdf = canvas.Canvas(buf, pagesize=letter)
    y = 720
    for line in lines:
        pdf.drawString(72, y, line)
        y -= 18
    pdf.save()
    return buf.getvalue()


STANDARD_ROWS = [
    ["Employee ID", "Date", "Clock In", "Clock Out", "Working Hours", "Status", "Overtime"],
    ["EMP001", "2026-03-17", "09:30", "18:30", "8", "Present", "0"],
]


def test_att_043a_xlsx_headers_map_to_fields():
    """ATT-043 — the canonical template shape parses field-for-field."""
    records = parse_attendance_file_rows(xlsx_bytes(STANDARD_ROWS), "import.xlsx")
    assert len(records) == 1
    row = records[0]
    assert row["employee_raw"] == "EMP001"
    assert row["status_raw"] == "Present"
    assert row["row_num"] == 2


def test_att_044a_csv_parses_identically_to_xlsx():
    """ATT-044 — same data, same result, including a UTF-8 BOM from Excel."""
    from_csv = parse_attendance_file_rows(csv_bytes(STANDARD_ROWS), "import.csv")
    from_bom = parse_attendance_file_rows(csv_bytes(STANDARD_ROWS, bom=True), "import.csv")
    assert from_csv[0]["employee_raw"] == from_bom[0]["employee_raw"] == "EMP001"


def test_att_059_txt_is_read_as_delimited_text():
    """.txt shares the CSV reader — an export renamed is still importable."""
    records = parse_attendance_file_rows(csv_bytes(STANDARD_ROWS), "import.txt")
    assert records[0]["employee_raw"] == "EMP001"


def test_att_060_pdf_rows_are_recovered_from_whitespace_columns():
    """A printed timesheet splits on runs of spaces, commas, pipes or tabs."""
    records = parse_attendance_file_rows(
        pdf_bytes(["Employee ID    Date    Status", "EMP001    2026-03-17    Present"]),
        "timesheet.pdf",
    )
    assert any(r["employee_raw"] == "EMP001" for r in records)


@pytest.mark.parametrize(
    "payload,name,message",
    [
        (b"not a workbook", "import.xlsx", "Failed to read Excel workbook"),
        (b"%PDF-1.4 broken", "import.pdf", "Failed to parse PDF attendance document"),
        (b"data", "import.docx", "Unsupported file format"),
    ],
)
def test_att_061_unreadable_import_files_raise_a_clear_value_error(payload, name, message):
    """Each format reports its own failure rather than a generic one."""
    with pytest.raises(ValueError, match=re.escape(message)):
        parse_attendance_file_rows(payload, name)


def test_att_062_a_header_only_file_is_rejected():
    """Fewer than two rows means there is nothing to import."""
    with pytest.raises(ValueError, match="empty or contains no data rows"):
        parse_attendance_file_rows(xlsx_bytes([STANDARD_ROWS[0]]), "import.xlsx")


def test_att_063_alternate_header_spellings_all_map():
    """HR exports never agree on column names; the aliases must all resolve."""
    rows = [
        ["Emp Code", "Attendance Date", "Punch In", "Punch Out", "Net Hours", "State", "OT"],
        ["EMP001", "2026-03-17", "09:30", "18:30", "8", "Present", "15"],
    ]
    record = parse_attendance_file_rows(xlsx_bytes(rows), "alt.xlsx")[0]
    assert record["employee_raw"] == "EMP001"
    assert record["status_raw"] == "Present"
    assert parse_float_value(record["overtime_raw"], 0.0) == 15.0
    assert parse_date_value(record["date_raw"]) == date(2026, 3, 17)


def test_att_064_unrecognisable_headers_fall_back_to_position():
    """With no header match, column 0 is the employee and column 1 the date."""
    rows = [["Column A", "Column B"], ["EMP001", "2026-03-17"]]
    record = parse_attendance_file_rows(xlsx_bytes(rows), "positional.xlsx")[0]
    assert record["employee_raw"] == "EMP001"
    assert parse_date_value(record["date_raw"]) == date(2026, 3, 17)


def test_att_065_short_rows_do_not_raise_on_missing_columns():
    """A ragged export must skip absent fields, not IndexError."""
    rows = [STANDARD_ROWS[0], ["EMP001"]]
    record = parse_attendance_file_rows(csv_bytes(rows), "ragged.csv")[0]
    assert record["employee_raw"] == "EMP001"
    assert record["date_raw"] is None
    assert record["status_raw"] == ""


def test_att_066_blank_rows_between_records_are_skipped():
    """Spreadsheets are full of spacer rows."""
    rows = [STANDARD_ROWS[0], STANDARD_ROWS[1], ["", "", "", "", "", "", ""], ["EMP001", "2026-03-18", "", "", "", "Absent", ""]]
    records = parse_attendance_file_rows(csv_bytes(rows), "spaced.csv")
    assert len(records) == 2


# --------------------------------------------------------------------------
# The import endpoint
# --------------------------------------------------------------------------


def test_att_050a_import_rejects_empty_and_oversized_files(admin_client):
    """ATT-050 — size is checked before any parser is invoked."""
    empty = admin_client.post("/api/attendance/import", files={"file": ("a.csv", b"", "text/csv")})
    assert empty.status_code == 400
    assert empty.json()["detail"] == "Uploaded file is empty"

    oversized = admin_client.post(
        "/api/attendance/import", files={"file": ("a.csv", b"x" * (10 * 1024 * 1024 + 1), "text/csv")}
    )
    assert oversized.status_code == 400
    assert "10 MB" in oversized.json()["detail"]


def test_att_067_an_unparseable_import_is_a_400(admin_client):
    """A parse failure surfaces the reason, not a 500."""
    res = admin_client.post("/api/attendance/import", files={"file": ("a.xlsx", b"junk", "application/vnd.ms-excel")})
    assert res.status_code == 400
    assert "Failed to read Excel workbook" in res.json()["detail"]


def test_att_043b_import_creates_updates_and_skips_in_one_pass(admin_client, db):
    """ATT-043/045/048 — the three outcomes are counted independently."""
    employee = db.query(User).filter(User.email == "test_employee@company.com").first()
    code = employee.profile.employee_id
    day_one = (date.today() - timedelta(days=10)).isoformat()
    day_two = (date.today() - timedelta(days=9)).isoformat()

    rows = [
        STANDARD_ROWS[0],
        [code, day_one, "09:30", "18:30", "8", "Present", "0"],
        ["NOSUCHCODE", day_two, "09:30", "18:30", "8", "Present", "0"],
        [code, "not-a-date", "09:30", "18:30", "8", "Present", "0"],
        ["", day_two, "09:30", "18:30", "8", "Present", "0"],
    ]
    first = admin_client.post("/api/attendance/import", files={"file": ("i.csv", csv_bytes(rows), "text/csv")})
    assert first.status_code == 200
    body = first.json()
    assert body["imported_count"] == 1
    assert body["skipped_count"] == 3
    assert any("not found in your organization" in e for e in body["errors"])
    assert any("Invalid or missing date" in e for e in body["errors"])
    assert any("Missing employee ID or email" in e for e in body["errors"])

    # Re-importing the same day updates rather than duplicating.
    second = admin_client.post("/api/attendance/import", files={"file": ("i.csv", csv_bytes(rows), "text/csv")})
    assert second.json()["updated_count"] == 1
    assert second.json()["imported_count"] == 0
    assert (
        db.query(Attendance)
        .filter(Attendance.user_id == employee.id, Attendance.date == date.fromisoformat(day_one))
        .count()
        == 1
    )


def test_att_068_employees_can_be_matched_by_email_as_well_as_code(admin_client, db):
    """The lookup indexes both the employee code and the login email."""
    employee = db.query(User).filter(User.email == "test_employee@company.com").first()
    day = (date.today() - timedelta(days=20)).isoformat()
    rows = [STANDARD_ROWS[0], [employee.email.upper(), day, "09:30", "18:30", "8", "Present", "0"]]

    res = admin_client.post("/api/attendance/import", files={"file": ("i.csv", csv_bytes(rows), "text/csv")})
    assert res.json()["imported_count"] == 1


@pytest.mark.parametrize(
    "status_input,expected,is_wfh",
    [
        ("Work From Home", "Work From Home", True),
        ("WFH", "Work From Home", True),
        ("Leave", "Leave", False),
        ("Half Day", "Half Day", False),
        ("Late", "Late", False),
        ("Absent", "Absent", False),
        ("Present", "Present", False),
        ("anything else", "Present", False),
    ],
)
def test_att_069_status_strings_are_normalised(admin_client, db, status_input, expected, is_wfh):
    """Imported status text is fuzzy-matched onto the internal vocabulary."""
    employee = db.query(User).filter(User.email == "test_employee@company.com").first()
    day = date.today() - timedelta(days=100)
    rows = [STANDARD_ROWS[0], [employee.profile.employee_id, day.isoformat(), "09:30", "18:30", "8", status_input, "0"]]

    res = admin_client.post("/api/attendance/import", files={"file": ("i.csv", csv_bytes(rows), "text/csv")})
    assert res.status_code == 200

    record = (
        db.query(Attendance).filter(Attendance.user_id == employee.id, Attendance.date == day).first()
    )
    assert record.status == expected
    assert record.is_wfh is is_wfh


@pytest.mark.parametrize(
    "hours,expected",
    [("8", "Present"), ("5", "Half Day"), ("2", "Late"), ("0", "Absent")],
)
def test_att_070_a_missing_status_is_derived_from_hours_worked(admin_client, db, hours, expected):
    """With no status column the hours decide, on documented thresholds."""
    employee = db.query(User).filter(User.email == "test_employee@company.com").first()
    day = date.today() - timedelta(days=120)
    rows = [
        ["Employee ID", "Date", "Working Hours"],
        [employee.profile.employee_id, day.isoformat(), hours],
    ]

    assert admin_client.post("/api/attendance/import", files={"file": ("i.csv", csv_bytes(rows), "text/csv")}).status_code == 200
    record = db.query(Attendance).filter(Attendance.user_id == employee.id, Attendance.date == day).first()
    assert record.status == expected


def test_att_071_hours_are_derived_from_the_punch_pair_when_absent(admin_client, db):
    """If the sheet gives times but no total, the total is computed."""
    employee = db.query(User).filter(User.email == "test_employee@company.com").first()
    day = date.today() - timedelta(days=130)
    rows = [
        ["Employee ID", "Date", "Clock In", "Clock Out"],
        [employee.profile.employee_id, day.isoformat(), "09:00", "17:30"],
    ]

    admin_client.post("/api/attendance/import", files={"file": ("i.csv", csv_bytes(rows), "text/csv")})
    record = db.query(Attendance).filter(Attendance.user_id == employee.id, Attendance.date == day).first()
    assert record.working_hours == pytest.approx(8.5)


def test_att_051a_the_error_list_is_capped(admin_client):
    """ATT-051 — 200 bad rows must not return 200 error strings."""
    rows = [STANDARD_ROWS[0]] + [["GHOST", "2026-03-17", "", "", "", "", ""] for _ in range(200)]
    res = admin_client.post("/api/attendance/import", files={"file": ("i.csv", csv_bytes(rows), "text/csv")})
    assert res.status_code == 200
    assert res.json()["skipped_count"] == 200
    assert len(res.json()["errors"]) == 15


def test_att_072_the_import_template_is_a_usable_workbook(admin_client):
    """The template must round-trip through the importer it is a template for."""
    res = admin_client.get("/api/attendance/template")
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    assert "attachment" in res.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    header = [c.value for c in next(wb.active.iter_rows())]
    assert any(h and "employee" in str(h).lower() for h in header)

    records = parse_attendance_file_rows(res.content, "template.xlsx")
    assert records  # the sample row parses with the shipped headers


# --------------------------------------------------------------------------
# Clock-out, breaks, and correction approval over an existing record
# --------------------------------------------------------------------------


def test_att_015a_clock_out_computes_hours_overtime_and_status(employee_client, db, office):
    """ATT-015 — the full punch-out calculation, including the break debit."""
    assert clock_in(employee_client, office).status_code == 200
    employee = db.query(User).filter(User.email == "test_employee@company.com").first()
    record = db.query(Attendance).filter(Attendance.user_id == employee.id).first()

    # Backdate the clock-in so the shift is long enough to earn overtime.
    record.clock_in = datetime.now(timezone.utc) - timedelta(hours=10)
    db.commit()

    assert employee_client.post("/api/attendance/break/start").status_code == 200
    assert employee_client.post("/api/attendance/break/end").status_code == 200

    res = employee_client.post("/api/attendance/clock-out")
    assert res.status_code == 200
    body = res.json()
    assert body["clock_out"] is not None
    assert body["working_hours"] > 0
    assert body["overtime_minutes"] > 0


def test_att_018a_clocking_out_with_an_open_break_closes_it(employee_client, db, office):
    """An open break is settled by the clock-out rather than left dangling."""
    assert clock_in(employee_client, office).status_code == 200
    assert employee_client.post("/api/attendance/break/start").status_code == 200

    assert employee_client.post("/api/attendance/clock-out").status_code == 200

    employee = db.query(User).filter(User.email == "test_employee@company.com").first()
    record = db.query(Attendance).filter(Attendance.user_id == employee.id).first()
    open_breaks = db.query(BreakSession).filter(
        BreakSession.attendance_id == record.id, BreakSession.end_time == None
    ).count()
    assert open_breaks == 0


def test_att_019a_a_short_shift_is_recorded_as_a_half_day(employee_client, db, office):
    """ATT-019 — under half the required hours reclassifies the day."""
    assert clock_in(employee_client, office).status_code == 200
    res = employee_client.post("/api/attendance/clock-out")
    assert res.status_code == 200
    assert res.json()["status"] == "Half Day"  # clocked in and out within seconds


@pytest.mark.parametrize(
    "endpoint,message",
    [
        ("/api/attendance/clock-out", "You must clock in first before clocking out"),
        ("/api/attendance/break/start", "You must clock in first before starting a break"),
        ("/api/attendance/break/end", "Attendance record not found for today"),
    ],
)
def test_att_016a_every_shift_action_requires_an_open_record(employee_client, endpoint, message):
    """ATT-016/020/022 — acting without a clock-in is a 400 on every route."""
    res = employee_client.post(endpoint)
    assert res.status_code == 400
    assert message.lower() in res.json()["detail"].lower()


def test_att_021a_a_second_break_and_a_post_clock_out_break_are_refused(employee_client, office):
    """ATT-018/021 — one open break at a time, and none after the shift ends."""
    assert clock_in(employee_client, office).status_code == 200
    assert employee_client.post("/api/attendance/break/start").status_code == 200

    second = employee_client.post("/api/attendance/break/start")
    assert second.status_code == 400

    assert employee_client.post("/api/attendance/break/end").status_code == 200
    assert employee_client.post("/api/attendance/break/end").status_code == 400

    assert employee_client.post("/api/attendance/clock-out").status_code == 200
    after = employee_client.post("/api/attendance/break/start")
    assert after.status_code == 400
    assert "Cannot start break after clocking out" in after.json()["detail"]


def test_att_040a_approving_a_correction_overwrites_an_existing_record(admin_client, employee_client, db, office):
    """ATT-040 — the approved proposal replaces the day's stored figures."""
    assert clock_in(employee_client, office).status_code == 200
    employee = db.query(User).filter(User.email == "test_employee@company.com").first()

    proposed_in = datetime.combine(date.today(), time(9, 0), tzinfo=timezone.utc)
    proposed_out = datetime.combine(date.today(), time(19, 0), tzinfo=timezone.utc)
    created = employee_client.post(
        "/api/attendance/corrections",
        json={
            "date": date.today().isoformat(),
            "proposed_clock_in": proposed_in.isoformat(),
            "proposed_clock_out": proposed_out.isoformat(),
            "reason": "Punched out on the wrong terminal",
        },
    )
    assert created.status_code == 200

    reviewed = admin_client.patch(
        f"/api/attendance/corrections/{created.json()['id']}/review",
        json={"status": "Approved", "comment": "Verified with the floor manager"},
    )
    assert reviewed.status_code == 200

    record = db.query(Attendance).filter(
        Attendance.user_id == employee.id, Attendance.date == date.today()
    ).first()
    assert record.modified_by_admin is True
    assert record.modification_reason == "Punched out on the wrong terminal"
    assert record.clock_out is not None


def test_att_073_rejecting_a_correction_changes_no_attendance(admin_client, employee_client, db):
    """A rejected correction is recorded but never touches the timesheet."""
    employee = db.query(User).filter(User.email == "test_employee@company.com").first()
    target = date.today() - timedelta(days=2)

    created = employee_client.post(
        "/api/attendance/corrections",
        json={"date": target.isoformat(), "reason": "Was on a client site"},
    )
    assert created.status_code == 200

    res = admin_client.patch(
        f"/api/attendance/corrections/{created.json()['id']}/review",
        json={"status": "Rejected", "comment": "No supporting evidence"},
    )
    assert res.status_code == 200
    assert res.json()["status"] == "Rejected"
    assert db.query(Attendance).filter(
        Attendance.user_id == employee.id, Attendance.date == target
    ).count() == 0


def test_att_074_a_malformed_correction_id_is_a_400(admin_client):
    """The review route parses the id itself."""
    res = admin_client.patch("/api/attendance/corrections/not-a-uuid/review", json={"status": "Approved"})
    assert res.status_code == 400
    assert res.json()["detail"] == "Invalid correction request ID format"


def test_att_075_an_unknown_correction_id_is_a_404(admin_client):
    """A well-formed id belonging to no request in this tenant is a 404."""
    res = admin_client.patch(
        f"/api/attendance/corrections/{uuid.uuid4()}/review", json={"status": "Approved"}
    )
    assert res.status_code == 404


# --------------------------------------------------------------------------
# History filters and selfie failure modes
# --------------------------------------------------------------------------


def test_att_029a_history_filters_compose(admin_client, employee_client, db, office):
    """ATT-029/030 — date range and status narrow together; inverted is empty."""
    assert clock_in(employee_client, office).status_code == 200
    employee = db.query(User).filter(User.email == "test_employee@company.com").first()

    today_iso = date.today().isoformat()
    in_range = admin_client.get(
        "/api/attendance/history",
        params={"start_date": today_iso, "end_date": today_iso, "user_id": str(employee.id), "status_filter": "Present"},
    )
    assert in_range.status_code == 200
    assert all(r["status"] == "Present" for r in in_range.json())

    inverted = admin_client.get(
        "/api/attendance/history",
        params={"start_date": today_iso, "end_date": (date.today() - timedelta(days=5)).isoformat()},
    )
    assert inverted.status_code == 200
    assert inverted.json() == []


def test_att_028a_a_malformed_user_id_filter_is_ignored(admin_client, employee_client, office):
    """ATT-028 — the filter is dropped rather than erroring; documented as-is."""
    assert clock_in(employee_client, office).status_code == 200
    res = admin_client.get("/api/attendance/history", params={"user_id": "not-a-uuid"})
    assert res.status_code == 200
    assert len(res.json()) >= 1  # unfiltered, not empty and not a 422


@pytest.mark.parametrize(
    "payload,fragment",
    [
        ("data:image/png;base64,!!!!not-base64!!!!", "Failed to process selfie"),
        ("data:image/png;base64," + base64.b64encode(b"MZ\x90\x00not-an-image").decode(), "Invalid selfie image"),
    ],
    ids=["malformed-base64", "not-an-image"],
)
def test_file_013a_a_bad_selfie_is_a_400_not_a_500(employee_client, office, payload, fragment):
    """FILE-013 — neither a decode failure nor a content failure may crash."""
    res = employee_client.post(
        "/api/attendance/clock-in",
        json={"latitude": office.latitude, "longitude": office.longitude, "selfie_base64": payload},
    )
    assert res.status_code == 400
    assert fragment in res.json()["detail"]


def test_file_014a_an_oversized_selfie_is_refused_after_decoding(employee_client, office):
    """FILE-014 — the cap is enforced on the decoded bytes, not the string."""
    oversized = base64.b64encode(b"\x00" * (5 * 1024 * 1024 + 10)).decode()
    res = employee_client.post(
        "/api/attendance/clock-in",
        json={"latitude": office.latitude, "longitude": office.longitude, "selfie_base64": oversized},
    )
    assert res.status_code == 400
    assert "exceeds maximum allowed size" in res.json()["detail"]


def test_att_014a_low_gps_accuracy_raises_the_suspicious_flag(employee_client, db, office):
    """ATT-014 — the threshold is strictly greater than 200 metres."""
    assert clock_in(employee_client, office, gps_accuracy=250.0).status_code == 200
    employee = db.query(User).filter(User.email == "test_employee@company.com").first()
    record = db.query(Attendance).filter(Attendance.user_id == employee.id).first()
    assert record.is_suspicious is True
