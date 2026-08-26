import io
import csv
from datetime import date, datetime, timezone
import pytest
from openpyxl import Workbook, load_workbook
import pypdf

from app.models.models import User, Attendance, EmployeeProfile, Department, Organization

def test_download_attendance_template(admin_client):
    response = admin_client.get("/api/attendance/template")
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers.get("content-type", "")
    
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) >= 2
    assert "Employee ID / Email" in rows[0]
    assert "Date (YYYY-MM-DD)" in rows[0]
    assert "Working Hours" in rows[0]

def test_attendance_excel_import(db, admin_client):
    admin = db.query(User).filter(User.email == "test_admin@company.com").first()
    org_id = admin.organization_id

    # Ensure test employee profile exists
    emp = db.query(User).filter(User.email == "emp_import_test@company.com").first()
    if not emp:
        emp = User(
            email="emp_import_test@company.com",
            hashed_password="hash",
            role="Employee",
            organization_id=org_id,
            is_active=True
        )
        db.add(emp)
        db.flush()

        profile = EmployeeProfile(
            organization_id=org_id,
            user_id=emp.id,
            employee_id="EMP-IMP-001",
            first_name="Import",
            last_name="Tester",
            hourly_rate=750.0,
            base_salary=120000.0
        )
        db.add(profile)
        db.commit()

    # Create Excel sheet in memory
    wb = Workbook()
    ws = wb.active
    ws.append(["Employee ID", "Date", "Clock In", "Clock Out", "Working Hours", "Status", "Overtime"])
    ws.append(["EMP-IMP-001", "2026-08-10", "09:30", "18:30", 8.0, "Present", 0])
    ws.append(["emp_import_test@company.com", "2026-08-11", "09:30", "19:30", 9.0, "Present", 60])
    ws.append(["NON_EXISTENT_EMP", "2026-08-11", "09:30", "18:30", 8.0, "Present", 0])

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    response = admin_client.post(
        "/api/attendance/import",
        files={"file": ("test_import.xlsx", excel_buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    assert response.status_code == 200
    data = response.json()
    assert data["total_rows"] == 3
    assert data["imported_count"] + data["updated_count"] == 2
    assert data["skipped_count"] == 1
    assert len(data["errors"]) >= 1
    assert "NON_EXISTENT_EMP" in data["errors"][0]

def test_attendance_csv_import(db, admin_client):
    admin = db.query(User).filter(User.email == "test_admin@company.com").first()
    org_id = admin.organization_id

    emp = db.query(User).filter(User.email == "emp_csv_test@company.com").first()
    if not emp:
        emp = User(
            email="emp_csv_test@company.com",
            hashed_password="hash",
            role="Employee",
            organization_id=org_id,
            is_active=True
        )
        db.add(emp)
        db.flush()

        profile = EmployeeProfile(
            organization_id=org_id,
            user_id=emp.id,
            employee_id="EMP-CSV-002",
            first_name="CSV",
            last_name="Tester",
            hourly_rate=600.0,
            base_salary=90000.0
        )
        db.add(profile)
        db.commit()

    csv_content = (
        "Employee ID,Date,Clock In,Clock Out,Working Hours,Status,Overtime\n"
        "EMP-CSV-002,2026-08-15,09:30,18:00,8.0,Work From Home,0\n"
        "EMP-CSV-002,2026-08-16,09:30,14:00,4.0,Half Day,0\n"
    )

    response = admin_client.post(
        "/api/attendance/import",
        files={"file": ("attendance.csv", csv_content.encode("utf-8"), "text/csv")}
    )
    assert response.status_code == 200
    data = response.json()
    assert data["imported_count"] + data["updated_count"] == 2
    assert data["skipped_count"] == 0

def test_attendance_import_unsupported_file(admin_client):
    response = admin_client.post(
        "/api/attendance/import",
        files={"file": ("image.png", b"\x89PNG\r\n\x1a\n", "image/png")}
    )
    assert response.status_code == 400
    assert "Unsupported file format" in response.json()["detail"]

def test_payroll_report_contains_user_id(admin_client):
    response = admin_client.get(
        "/api/reports/payroll?start_date=2026-08-01&end_date=2026-08-31"
    )
    assert response.status_code == 200
    data = response.json()
    assert isinstance(data, list)
    assert len(data) > 0
    assert "user_id" in data[0]
    assert "designation" in data[0]
    assert "total_salary" in data[0]

def test_individual_payslip_pdf_generation(db, admin_client):
    emp = db.query(User).filter(User.email == "test_employee@company.com").first()
    assert emp is not None

    response = admin_client.get(
        f"/api/reports/payslip/{emp.id}/pdf?start_date=2026-08-01&end_date=2026-08-31"
    )
    assert response.status_code == 200
    assert response.headers.get("content-type") == "application/pdf"
    assert "attachment; filename=payslip_" in response.headers.get("content-disposition", "")
    assert len(response.content) > 1000

    # Validate PDF structure with pypdf
    pdf_reader = pypdf.PdfReader(io.BytesIO(response.content))
    assert len(pdf_reader.pages) >= 1
    extracted_text = pdf_reader.pages[0].extract_text()
    assert "SALARY PAYSLIP" in extracted_text
    assert "NET SALARY PAYABLE" in extracted_text

def test_individual_payslip_pdf_invalid_user_id(admin_client):
    response = admin_client.get(
        "/api/reports/payslip/invalid-uuid/pdf"
    )
    assert response.status_code == 400
    assert "Invalid employee user ID" in response.json()["detail"]
