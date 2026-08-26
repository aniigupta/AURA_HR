import csv
import io
import uuid
from datetime import date, datetime, timedelta
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from app.core.database import get_db
from app.core.security import RoleChecker
from app.models.models import User, Attendance, EmployeeProfile, OfficeSetting
from app.core.utils import get_day_status_for_employee
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

router = APIRouter(prefix="/reports", tags=["Reports & Exports"])

admin_required = RoleChecker(["Admin"])

def get_report_data(
    db: Session,
    organization_id: Any,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    department_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    employee_id: Optional[str] = None
) -> List[dict]:
    query = db.query(Attendance).options(
        joinedload(Attendance.user).joinedload(User.profile).joinedload(EmployeeProfile.department)
    ).filter(Attendance.organization_id == organization_id)

    if start_date:
        query = query.filter(Attendance.date >= start_date)
    if end_date:
        query = query.filter(Attendance.date <= end_date)
    if department_id or employee_id:
        query = query.join(Attendance.user).join(User.profile)
        if department_id:
            query = query.filter(EmployeeProfile.department_id == department_id)
        if employee_id:
            query = query.filter(EmployeeProfile.id == employee_id)
    if status_filter:
        query = query.filter(Attendance.status == status_filter)

    attendances = query.order_by(Attendance.date.desc()).all()

    report_list = []
    for att in attendances:
        profile = att.user.profile if att.user else None
        dept_name = profile.department.name if profile and profile.department else "N/A"
        emp_name = f"{profile.first_name} {profile.last_name}" if profile else (att.user.email if att.user else "N/A")
        emp_code = profile.employee_id if profile else "N/A"
        
        hourly_rate = round(profile.hourly_rate, 2) if profile and profile.hourly_rate else 0.0
        earned_salary = round(att.working_hours * hourly_rate, 2)

        report_list.append({
            "date": att.date.isoformat(),
            "employee_id": emp_code,
            "name": emp_name,
            "email": att.user.email if att.user else "N/A",
            "department": dept_name,
            "clock_in": att.clock_in.strftime("%H:%M:%S") if att.clock_in else "N/A",
            "clock_out": att.clock_out.strftime("%H:%M:%S") if att.clock_out else "N/A",
            "working_hours": round(att.working_hours, 2),
            "break_duration": round(att.break_duration, 2),
            "hourly_rate": hourly_rate,
            "earned_salary": earned_salary,
            "status": att.status,
            "late_minutes": att.late_minutes,
            "overtime_minutes": att.overtime_minutes
        })
    return report_list

@router.get("/summary")
def get_reports_summary(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    department_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    employee_id: Optional[str] = None,
    db: Session = Depends(get_db),
    admin_user: User = Depends(admin_required)
):
    return get_report_data(db, admin_user.organization_id, start_date, end_date, department_id, status_filter, employee_id)

@router.get("/export/csv")
def export_csv(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    department_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    employee_id: Optional[str] = None,
    db: Session = Depends(get_db),
    admin_user: User = Depends(admin_required)
):
    data = get_report_data(db, admin_user.organization_id, start_date, end_date, department_id, status_filter, employee_id)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        "Date", "Employee ID", "Name", "Email", "Department", 
        "Clock In", "Clock Out", "Working Hours", "Break Hours", 
        "Hourly Rate (INR)", "Earned Daily Salary (INR)",
        "Status", "Late (Mins)", "Overtime (Mins)"
    ])
    
    for row in data:
        writer.writerow([
            row["date"], row["employee_id"], row["name"], row["email"], row["department"],
            row["clock_in"], row["clock_out"], row["working_hours"], row["break_duration"],
            f"INR {row['hourly_rate']:.2f}", f"INR {row['earned_salary']:.2f}",
            row["status"], row["late_minutes"], row["overtime_minutes"]
        ])
        
    output.seek(0)
    filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/excel")
def export_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    department_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    employee_id: Optional[str] = None,
    db: Session = Depends(get_db),
    admin_user: User = Depends(admin_required)
):
    data = get_report_data(db, admin_user.organization_id, start_date, end_date, department_id, status_filter, employee_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    headers = [
        "Date", "Employee ID", "Name", "Email", "Department", 
        "Clock In", "Clock Out", "Working Hours", "Break Hours", 
        "Hourly Rate (INR)", "Earned Salary (INR)",
        "Status", "Late (Mins)", "Overtime (Mins)"
    ]
    
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for row in data:
        ws.append([
            row["date"], row["employee_id"], row["name"], row["email"], row["department"],
            row["clock_in"], row["clock_out"], row["working_hours"], row["break_duration"],
            row["hourly_rate"], row["earned_salary"],
            row["status"], row["late_minutes"], row["overtime_minutes"]
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    for r_idx in range(2, ws.max_row + 1):
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.alignment = center_align if c_idx in [1, 2, 6, 7, 8, 9, 10, 11, 12, 13, 14] else left_align

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/pdf")
def export_pdf(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    department_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    employee_id: Optional[str] = None,
    db: Session = Depends(get_db),
    admin_user: User = Depends(admin_required)
):
    data = get_report_data(db, admin_user.organization_id, start_date, end_date, department_id, status_filter, employee_id)
    
    pdf_buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=letter,
        rightMargin=20,
        leftMargin=20,
        topMargin=30,
        bottomMargin=30
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="TitleStyle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#1F2937"),
        spaceAfter=15
    )
    
    cell_style = ParagraphStyle(
        name="CellStyle",
        fontName="Helvetica",
        fontSize=8,
        leading=10
    )
    
    cell_header_style = ParagraphStyle(
        name="CellHeaderStyle",
        fontName="Helvetica-Bold",
        fontSize=8,
        textColor=colors.white,
        leading=10
    )

    story = []
    org_name = admin_user.organization.name if admin_user.organization else "AuraWork"
    story.append(Paragraph(f"{org_name} - Workforce Attendance & Salary Report (INR)", title_style))
    story.append(Spacer(1, 10))
    
    headers = [
        Paragraph("Date", cell_header_style),
        Paragraph("ID", cell_header_style),
        Paragraph("Name", cell_header_style),
        Paragraph("Department", cell_header_style),
        Paragraph("Clock In", cell_header_style),
        Paragraph("Clock Out", cell_header_style),
        Paragraph("Hours", cell_header_style),
        Paragraph("Rate (INR)", cell_header_style),
        Paragraph("Earned (INR)", cell_header_style),
        Paragraph("Status", cell_header_style)
    ]
    
    table_data = [headers]
    for row in data:
        table_data.append([
            Paragraph(row["date"], cell_style),
            Paragraph(row["employee_id"], cell_style),
            Paragraph(row["name"], cell_style),
            Paragraph(row["department"], cell_style),
            Paragraph(row["clock_in"], cell_style),
            Paragraph(row["clock_out"], cell_style),
            Paragraph(str(row["working_hours"]), cell_style),
            Paragraph(f"INR {row['hourly_rate']:.2f}", cell_style),
            Paragraph(f"INR {row['earned_salary']:.2f}", cell_style),
            Paragraph(row["status"], cell_style)
        ])
        
    col_widths = [60, 40, 85, 75, 50, 50, 40, 55, 60, 55]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F2937")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
        ('TOPPADDING', (0,1), (-1,-1), 4),
        ('BOTTOMPADDING', (0,1), (-1,-1), 4),
    ])
    t.setStyle(t_style)
    story.append(t)
    
    doc.build(story)
    pdf_buffer.seek(0)
    
    filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def get_payroll_summary_data(db: Session, organization_id: Any, start_date: date, end_date: date) -> List[dict]:
    settings = db.query(OfficeSetting).filter(OfficeSetting.organization_id == organization_id).first()
    if not settings:
        settings = OfficeSetting(organization_id=organization_id)
        db.add(settings)
        db.commit()
        db.refresh(settings)

    employees = db.query(EmployeeProfile).options(
        joinedload(EmployeeProfile.department),
        joinedload(EmployeeProfile.user)
    ).join(User).filter(
        User.organization_id == organization_id,
        User.role == "Employee",
        User.is_active == True
    ).all()

    all_attendances = db.query(Attendance).filter(
        Attendance.organization_id == organization_id,
        Attendance.date >= start_date,
        Attendance.date <= end_date
    ).all()

    user_attendance_map: Dict[Any, Dict[date, Attendance]] = {}
    for att in all_attendances:
        if att.user_id not in user_attendance_map:
            user_attendance_map[att.user_id] = {}
        user_attendance_map[att.user_id][att.date] = att

    payroll_list = []
    delta = end_date - start_date
    date_list = [start_date + timedelta(days=i) for i in range(delta.days + 1)]
    
    for emp in employees:
        user_id = emp.user_id
        emp_name = f"{emp.first_name} {emp.last_name}"
        emp_code = emp.employee_id
        dept_name = emp.department.name if emp.department else "N/A"
        rate = emp.hourly_rate or 0.0
        base = emp.base_salary or 0.0
        
        emp_attendance_map = user_attendance_map.get(user_id, {})
        
        present_days = 0
        wfh_days = 0
        leave_days = 0
        absent_days = 0
        half_days = 0
        late_days = 0
        total_hours = 0.0
        overtime_hours = 0.0
        
        for d in date_list:
            if d in emp_attendance_map:
                att = emp_attendance_map[d]
                total_hours += att.working_hours
                overtime_hours += (att.overtime_minutes / 60.0)
                
                status_str = att.status
                if status_str == "Present":
                    present_days += 1
                elif status_str == "Late":
                    present_days += 1
                    late_days += 1
                elif status_str == "Half Day":
                    present_days += 1
                    half_days += 1
                elif status_str == "Work From Home":
                    wfh_days += 1
                elif status_str == "Leave":
                    leave_days += 1
                elif status_str == "Absent":
                    absent_days += 1
            else:
                fallback_status = get_day_status_for_employee(db, user_id, d, emp, settings)
                if fallback_status == "Leave":
                    leave_days += 1
                elif fallback_status == "Absent":
                    absent_days += 1
                elif fallback_status == "Work From Home":
                    wfh_days += 1
                
        working_salary = round(total_hours * rate, 2)
        overtime_pay = round(overtime_hours * rate * 1.5, 2)
        total_salary = round(working_salary + overtime_pay, 2)

        payroll_list.append({
            "user_id": str(user_id),
            "employee_id": emp_code,
            "name": emp_name,
            "department": dept_name,
            "designation": emp.designation or "Staff",
            "hourly_rate": round(rate, 2),
            "base_salary": round(base, 2),
            "present_days": present_days,
            "wfh_days": wfh_days,
            "leave_days": leave_days,
            "absent_days": absent_days,
            "half_days": half_days,
            "late_days": late_days,
            "total_hours": round(total_hours, 2),
            "overtime_hours": round(overtime_hours, 2),
            "working_salary": working_salary,
            "overtime_pay": overtime_pay,
            "total_salary": total_salary,
            "total_days_in_period": len(date_list)
        })
        
    return payroll_list

@router.get("/payroll")
def get_payroll_report(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    admin_user: User = Depends(admin_required)
):
    today = date.today()
    s_date = start_date or today.replace(day=1)
    e_date = end_date or today
    return get_payroll_summary_data(db, admin_user.organization_id, s_date, e_date)

@router.get("/export/payroll")
def export_payroll_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    admin_user: User = Depends(admin_required)
):
    today = date.today()
    s_date = start_date or today.replace(day=1)
    e_date = end_date or today
    data = get_payroll_summary_data(db, admin_user.organization_id, s_date, e_date)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Report"
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    headers = [
        "Employee ID", "Name", "Department", "Hourly Rate (INR)", "Present Days", 
        "WFH Days", "Leave Days", "Absent Days", "Half Days", 
        "Late Days", "Total Hours", "Overtime Hours", 
        "Working Salary (INR)", "Overtime Pay (INR)", "Total Salary (INR)"
    ]
    
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
    for row in data:
        ws.append([
            row["employee_id"], row["name"], row["department"], row["hourly_rate"],
            row["present_days"], row["wfh_days"], row["leave_days"], row["absent_days"],
            row["half_days"], row["late_days"], row["total_hours"], row["overtime_hours"],
            row["working_salary"], row["overtime_pay"], row["total_salary"]
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    for r_idx in range(2, ws.max_row + 1):
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.alignment = center_align if c_idx not in [2, 3] else left_align
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"payroll_salary_report_{s_date}_{e_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/payroll/pdf")
def export_payroll_pdf(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    admin_user: User = Depends(admin_required)
):
    today = date.today()
    s_date = start_date or today.replace(day=1)
    e_date = end_date or today
    data = get_payroll_summary_data(db, admin_user.organization_id, s_date, e_date)
    
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=letter,
        rightMargin=20,
        leftMargin=20,
        topMargin=30,
        bottomMargin=30
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="PayrollTitleStyle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#1F2937"),
        spaceAfter=15
    )
    
    cell_style = ParagraphStyle(
        name="PayrollCellStyle",
        fontName="Helvetica",
        fontSize=8,
        leading=10
    )
    
    cell_header_style = ParagraphStyle(
        name="PayrollHeaderStyle",
        fontName="Helvetica-Bold",
        fontSize=8,
        textColor=colors.white,
        leading=10
    )

    story = []
    org_name = admin_user.organization.name if admin_user.organization else "AuraWork"
    story.append(Paragraph(f"{org_name} - Corporate Employee Payroll & Salary Statement (INR ₹)", title_style))
    story.append(Spacer(1, 10))
    
    headers = [
        Paragraph("ID", cell_header_style),
        Paragraph("Name", cell_header_style),
        Paragraph("Dept", cell_header_style),
        Paragraph("Rate/hr", cell_header_style),
        Paragraph("Present", cell_header_style),
        Paragraph("Leave", cell_header_style),
        Paragraph("Absent", cell_header_style),
        Paragraph("Hours", cell_header_style),
        Paragraph("OT Hrs", cell_header_style),
        Paragraph("Work Pay", cell_header_style),
        Paragraph("OT Pay", cell_header_style),
        Paragraph("Total Pay (INR)", cell_header_style)
    ]
    
    table_data = [headers]
    for row in data:
        table_data.append([
            Paragraph(row["employee_id"], cell_style),
            Paragraph(row["name"], cell_style),
            Paragraph(row["department"], cell_style),
            Paragraph(f"INR {row['hourly_rate']:.2f}", cell_style),
            Paragraph(f"{row['present_days']}d", cell_style),
            Paragraph(f"{row['leave_days']}d", cell_style),
            Paragraph(f"{row['absent_days']}d", cell_style),
            Paragraph(f"{row['total_hours']}h", cell_style),
            Paragraph(f"{row['overtime_hours']}h", cell_style),
            Paragraph(f"INR {row['working_salary']:.2f}", cell_style),
            Paragraph(f"INR {row['overtime_pay']:.2f}", cell_style),
            Paragraph(f"INR {row['total_salary']:.2f}", cell_style)
        ])
        
    col_widths = [45, 80, 60, 45, 40, 35, 35, 40, 40, 55, 50, 60]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F2937")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
        ('TOPPADDING', (0,1), (-1,-1), 4),
        ('BOTTOMPADDING', (0,1), (-1,-1), 4),
    ])
    t.setStyle(t_style)
    story.append(t)
    
    doc.build(story)
    pdf_buffer.seek(0)
    
    filename = f"payroll_salary_report_{s_date}_{e_date}.pdf"
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# --- Individual Employee Salary Slip (Payslip PDF) Generator ---

def generate_individual_payslip_pdf(
    db: Session,
    organization: Any,
    user_id: Any,
    start_date: date,
    end_date: date
) -> io.BytesIO:
    """
    Generates a formal corporate Salary Payslip PDF customized to employee attendance,
    hours worked, overtime, and statutory deductions for the pay period.
    """
    user = db.query(User).options(
        joinedload(User.profile).joinedload(EmployeeProfile.department)
    ).filter(
        User.id == user_id,
        User.organization_id == organization.id
    ).first()

    if not user or not user.profile:
        raise HTTPException(status_code=404, detail="Employee profile not found")

    profile = user.profile
    dept_name = profile.department.name if profile.department else "General"
    emp_name = f"{profile.first_name} {profile.last_name}"
    emp_code = profile.employee_id
    designation = profile.designation or "Staff"
    rate = profile.hourly_rate or 0.0
    base_salary = profile.base_salary or 0.0

    setting = db.query(OfficeSetting).filter(OfficeSetting.organization_id == organization.id).first()
    if not setting:
        setting = OfficeSetting(organization_id=organization.id)

    attendances = db.query(Attendance).filter(
        Attendance.organization_id == organization.id,
        Attendance.user_id == user_id,
        Attendance.date >= start_date,
        Attendance.date <= end_date
    ).all()

    att_map = {att.date: att for att in attendances}
    delta = end_date - start_date
    date_list = [start_date + timedelta(days=i) for i in range(delta.days + 1)]

    present_days = 0
    wfh_days = 0
    leave_days = 0
    absent_days = 0
    half_days = 0
    late_days = 0
    total_hours = 0.0
    overtime_hours = 0.0

    for d in date_list:
        if d in att_map:
            att = att_map[d]
            total_hours += att.working_hours
            overtime_hours += (att.overtime_minutes / 60.0)
            if att.status == "Present":
                present_days += 1
            elif att.status == "Late":
                present_days += 1
                late_days += 1
            elif att.status == "Half Day":
                present_days += 1
                half_days += 1
            elif att.status == "Work From Home":
                wfh_days += 1
            elif att.status == "Leave":
                leave_days += 1
            elif att.status == "Absent":
                absent_days += 1
        else:
            fb = get_day_status_for_employee(db, user_id, d, profile, setting)
            if fb == "Leave":
                leave_days += 1
            elif fb == "Absent":
                absent_days += 1
            elif fb == "Work From Home":
                wfh_days += 1

    working_salary = round(total_hours * rate, 2) if rate > 0 else base_salary
    effective_ot_rate = rate if rate > 0 else (base_salary / 160.0 if base_salary > 0 else 500.0)
    overtime_pay = round(overtime_hours * effective_ot_rate * 1.5, 2)
    gross_earnings = round(working_salary + overtime_pay, 2)

    provident_fund = round(min(gross_earnings * 0.12, 1800.0), 2) if gross_earnings > 15000 else 0.0
    professional_tax = 200.0 if gross_earnings > 10000 else 0.0
    total_deductions = round(provident_fund + professional_tax, 2)
    net_payable = round(max(gross_earnings - total_deductions, 0.0), 2)

    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=letter,
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25
    )

    styles = getSampleStyleSheet()
    primary_color = colors.HexColor("#1E1B4B")
    accent_color = colors.HexColor("#4F46E5")
    border_color = colors.HexColor("#CBD5E1")
    light_bg = colors.HexColor("#F8FAFC")

    title_style = ParagraphStyle(
        name="SlipTitleStyle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=primary_color,
        spaceAfter=3
    )

    subtitle_style = ParagraphStyle(
        name="SlipSubtitleStyle",
        fontName="Helvetica-Bold",
        fontSize=11,
        textColor=accent_color,
        spaceAfter=12
    )

    section_header = ParagraphStyle(
        name="SlipSecHeader",
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.white,
        leading=11
    )

    label_style = ParagraphStyle(
        name="SlipLabelStyle",
        fontName="Helvetica-Bold",
        fontSize=8,
        textColor=colors.HexColor("#475569"),
        leading=10
    )

    val_style = ParagraphStyle(
        name="SlipValStyle",
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.HexColor("#0F172A"),
        leading=10
    )

    story = []
    org_name = organization.name if organization else "AuraWork Enterprises"
    period_str = f"{start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    month_name = start_date.strftime("%B %Y")

    story.append(Paragraph(f"{org_name}", title_style))
    story.append(Paragraph(f"CORPORATE SALARY PAYSLIP — {month_name.upper()} (INR ₹)", subtitle_style))
    story.append(Spacer(1, 4))

    # Employee Information Grid
    emp_info_data = [
        [
            Paragraph("Employee Name", label_style), Paragraph(emp_name, val_style),
            Paragraph("Pay Period", label_style), Paragraph(period_str, val_style)
        ],
        [
            Paragraph("Employee ID", label_style), Paragraph(emp_code, val_style),
            Paragraph("Designation", label_style), Paragraph(designation, val_style)
        ],
        [
            Paragraph("Department", label_style), Paragraph(dept_name, val_style),
            Paragraph("Joining Date", label_style), Paragraph(str(profile.join_date or "N/A"), val_style)
        ],
        [
            Paragraph("Hourly Rate", label_style), Paragraph(f"INR {rate:.2f} / hr", val_style),
            Paragraph("Payment Mode", label_style), Paragraph("Bank Transfer (NEFT/RTGS)", val_style)
        ]
    ]

    t_emp = Table(emp_info_data, colWidths=[90, 190, 90, 190])
    t_emp.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), light_bg),
        ('BOX', (0,0), (-1,-1), 1, border_color),
        ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(t_emp)
    story.append(Spacer(1, 10))

    # Attendance Metrics Ribbon
    att_summary_data = [
        [
            Paragraph("<b>Total Days</b>", label_style),
            Paragraph("<b>Present</b>", label_style),
            Paragraph("<b>WFH</b>", label_style),
            Paragraph("<b>Paid Leaves</b>", label_style),
            Paragraph("<b>Absences</b>", label_style),
            Paragraph("<b>Hours Logged</b>", label_style),
            Paragraph("<b>Overtime</b>", label_style),
        ],
        [
            Paragraph(f"{len(date_list)}", val_style),
            Paragraph(f"{present_days} d", val_style),
            Paragraph(f"{wfh_days} d", val_style),
            Paragraph(f"{leave_days} d", val_style),
            Paragraph(f"{absent_days} d", val_style),
            Paragraph(f"{total_hours:.2f} hrs", val_style),
            Paragraph(f"{overtime_hours:.2f} hrs", val_style),
        ]
    ]
    t_att = Table(att_summary_data, colWidths=[80, 80, 80, 80, 80, 80, 80])
    t_att.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#EEF2F6")),
        ('BACKGROUND', (0,1), (-1,1), colors.white),
        ('BOX', (0,0), (-1,-1), 1, border_color),
        ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t_att)
    story.append(Spacer(1, 12))

    # Earnings & Deductions Breakdown Table
    breakdown_data = [
        [Paragraph("EARNINGS", section_header), Paragraph("AMOUNT (INR)", section_header), Paragraph("DEDUCTIONS", section_header), Paragraph("AMOUNT (INR)", section_header)],
        [Paragraph("Working Hours Base Salary", val_style), Paragraph(f"INR {working_salary:.2f}", val_style), Paragraph("Employee Provident Fund (PF)", val_style), Paragraph(f"INR {provident_fund:.2f}", val_style)],
        [Paragraph(f"Overtime Payout ({overtime_hours:.1f} hrs @ 1.5x)", val_style), Paragraph(f"INR {overtime_pay:.2f}", val_style), Paragraph("Professional Tax (PT)", val_style), Paragraph(f"INR {professional_tax:.2f}", val_style)],
        [Paragraph("Performance & Attendance Allowance", val_style), Paragraph("INR 0.00", val_style), Paragraph("Income Tax / TDS Deduction", val_style), Paragraph("INR 0.00", val_style)],
        [Paragraph("<b>TOTAL GROSS EARNINGS</b>", label_style), Paragraph(f"<b>INR {gross_earnings:.2f}</b>", label_style), Paragraph("<b>TOTAL DEDUCTIONS</b>", label_style), Paragraph(f"<b>INR {total_deductions:.2f}</b>", label_style)]
    ]

    t_breakdown = Table(breakdown_data, colWidths=[190, 90, 190, 90])
    t_breakdown.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('BACKGROUND', (0,1), (-1,-2), colors.white),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#F1F5F9")),
        ('BOX', (0,0), (-1,-1), 1, primary_color),
        ('INNERGRID', (0,0), (-1,-1), 0.5, border_color),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(t_breakdown)
    story.append(Spacer(1, 12))

    # Net Salary Highlight Banner
    net_data = [[
        Paragraph(f"<b>NET SALARY PAYABLE: INR ₹{net_payable:,.2f}</b>", ParagraphStyle(
            name="NetPayStyle", fontName="Helvetica-Bold", fontSize=13, textColor=colors.white, alignment=1
        ))
    ]]
    t_net = Table(net_data, colWidths=[560])
    t_net.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), accent_color),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    ]))
    story.append(t_net)
    story.append(Spacer(1, 20))

    # Signatory and Notes
    sig_data = [
        [
            Paragraph("<i>Note: This is a system-generated salary slip and does not require a physical signature.</i><br/>"
                      f"Generated on {datetime.now().strftime('%d-%b-%Y %H:%M')} UTC | AuraHR Cloud", ParagraphStyle(
                name="NotesStyle", fontName="Helvetica", fontSize=7, textColor=colors.HexColor("#64748B"), leading=9
            )),
            Paragraph("<b>Authorized Signatory</b><br/><br/>_______________________<br/>HR & Finance Dept", ParagraphStyle(
                name="SigStyle", fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#1E293B"), alignment=2, leading=11
            ))
        ]
    ]
    t_sig = Table(sig_data, colWidths=[360, 200])
    t_sig.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(t_sig)

    doc.build(story)
    pdf_buffer.seek(0)
    return pdf_buffer

@router.get("/payslip/{user_id}/pdf")
def download_individual_payslip_pdf(
    user_id: str,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    admin_user: User = Depends(admin_required)
):
    """
    Downloads an official corporate Salary Payslip PDF for an individual employee.
    """
    try:
        uuid_user_id = uuid.UUID(user_id) if isinstance(user_id, str) else user_id
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid employee user ID format")

    today = date.today()
    s_date = start_date or today.replace(day=1)
    e_date = end_date or today

    pdf_buffer = generate_individual_payslip_pdf(
        db, admin_user.organization, uuid_user_id, s_date, e_date
    )

    filename = f"payslip_{user_id}_{s_date}_{e_date}.pdf"
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

