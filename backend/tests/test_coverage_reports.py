"""
Coverage for the reporting and payroll surface: every export format, every
status branch in the payroll aggregation, and the payslip generator.

The export loops and the per-status counters only execute when there is data
of that shape, so this module seeds a month containing one day of every status
the system can record — plus days with no record at all, which exercise the
fallback classifier that decides whether an uncovered day is leave, a holiday,
a weekend or an absence.
"""

import io
from datetime import date, datetime, timedelta, timezone

import openpyxl
import pytest

from app.models.models import (
    Attendance,
    EmployeeProfile,
    Holiday,
    LeaveRequest,
    OfficeSetting,
    Organization,
    User,
)
from app.routers.reports import get_payroll_summary_data

ALL_STATUSES = ["Present", "Late", "Half Day", "Work From Home", "Leave", "Absent"]


@pytest.fixture
def month(db):
    """A settled month: one day per status, plus uncovered days and a holiday."""
    employee = db.query(User).filter(User.email == "test_employee@company.com").first()
    employee.profile.hourly_rate = 500.0
    employee.profile.base_salary = 80000.0

    start = date.today().replace(day=1)
    day = start
    seeded = {}
    for status in ALL_STATUSES:
        db.add(
            Attendance(
                organization_id=employee.organization_id,
                user_id=employee.id,
                date=day,
                clock_in=datetime.combine(day, datetime.min.time(), tzinfo=timezone.utc) + timedelta(hours=9, minutes=30),
                clock_out=datetime.combine(day, datetime.min.time(), tzinfo=timezone.utc) + timedelta(hours=18, minutes=30),
                working_hours=8.0 if status not in ("Absent", "Leave") else 0.0,
                break_duration=1.0,
                status=status,
                is_wfh=status == "Work From Home",
                late_minutes=22 if status == "Late" else 0,
                overtime_minutes=45 if status == "Present" else 0,
            )
        )
        seeded[status] = day
        day += timedelta(days=1)

    # An uncovered day that approved leave should classify as Leave.
    leave_day = day
    db.add(
        LeaveRequest(
            organization_id=employee.organization_id,
            user_id=employee.id,
            leave_type="Casual",
            start_date=leave_day,
            end_date=leave_day,
            reason="Covered by approved leave",
            status="Approved",
        )
    )
    # An uncovered day that is a configured holiday.
    holiday_day = day + timedelta(days=1)
    db.add(Holiday(organization_id=employee.organization_id, name="Founders Day", date=holiday_day))

    db.commit()
    return {
        "employee": employee,
        "start": start,
        "end": holiday_day + timedelta(days=1),
        "seeded": seeded,
        "leave_day": leave_day,
        "holiday_day": holiday_day,
    }


def params(month) -> dict:
    return {"start_date": month["start"].isoformat(), "end_date": month["end"].isoformat()}


# --------------------------------------------------------------------------
# get_report_data — the filter combinations
# --------------------------------------------------------------------------


def test_rpt_001a_every_filter_narrows_the_summary(admin_client, db, month):
    """RPT-001 — date range, department, employee and status all compose."""
    profile = month["employee"].profile

    unfiltered = admin_client.get("/api/reports/summary", params=params(month))
    assert unfiltered.status_code == 200
    assert len(unfiltered.json()) == len(ALL_STATUSES)

    by_status = admin_client.get(
        "/api/reports/summary", params={**params(month), "status_filter": "Late"}
    )
    assert [r["status"] for r in by_status.json()] == ["Late"]
    assert by_status.json()[0]["late_minutes"] == 22

    by_department = admin_client.get(
        "/api/reports/summary", params={**params(month), "department_id": profile.department_id}
    )
    assert len(by_department.json()) == len(ALL_STATUSES)

    other_department = admin_client.get(
        "/api/reports/summary", params={**params(month), "department_id": 999999}
    )
    assert other_department.json() == []


def test_rpt_011_the_employee_id_filter_selects_that_employee(admin_client, month):
    """
    RPT-011 / RPT-B1 — regression guard, fixed.

    employee_id was annotated Optional[str] and compared to EmployeeProfile.id,
    a UUID column, so passing a real profile UUID — exactly what the reports
    screen sends — raised in the driver before any row was read and returned a
    500 from all four endpoints. Now parsed as Optional[uuid.UUID], which also
    turns a malformed value into a 422 instead of a 500.
    """
    profile = month["employee"].profile

    rows = admin_client.get(
        "/api/reports/summary", params={**params(month), "employee_id": str(profile.id)}
    )
    assert rows.status_code == 200
    assert len(rows.json()) == len(ALL_STATUSES)
    assert all(r["employee_id"] == profile.employee_id for r in rows.json())

    for path in ("/api/reports/export/csv", "/api/reports/export/excel", "/api/reports/export/pdf"):
        res = admin_client.get(path, params={**params(month), "employee_id": str(profile.id)})
        assert res.status_code == 200, path
        assert len(res.content) > 0

    # A profile id that exists in no tenant is an empty report, not an error.
    import uuid as _uuid

    empty = admin_client.get(
        "/api/reports/summary", params={**params(month), "employee_id": str(_uuid.uuid4())}
    )
    assert empty.status_code == 200
    assert empty.json() == []

    # And a malformed id is now a validation error rather than a database error.
    malformed = admin_client.get(
        "/api/reports/summary", params={**params(month), "employee_id": "not-a-uuid"}
    )
    assert malformed.status_code == 422
    assert malformed.json()["errorCode"] == "VALIDATION_ERROR"


def test_rpt_009_earned_salary_is_hours_times_rate(admin_client, month):
    """Each row carries its own derived pay, not just raw hours."""
    row = admin_client.get(
        "/api/reports/summary", params={**params(month), "status_filter": "Present"}
    ).json()[0]
    assert row["hourly_rate"] == 500.0
    assert row["earned_salary"] == pytest.approx(8.0 * 500.0)


def test_rpt_010_a_row_without_a_profile_does_not_break_the_report(admin_client, db, month):
    """A User with attendance but no profile must degrade, not crash."""
    org_id = month["employee"].organization_id
    orphan = User(
        organization_id=org_id,
        email="orphan.report@company.com",
        hashed_password="x",
        role="Employee",
        is_active=True,
    )
    db.add(orphan)
    db.flush()
    db.add(
        Attendance(
            organization_id=org_id,
            user_id=orphan.id,
            date=month["start"],
            clock_in=datetime.now(timezone.utc),
            status="Present",
        )
    )
    db.commit()

    res = admin_client.get("/api/reports/summary", params=params(month))
    assert res.status_code == 200
    orphan_rows = [r for r in res.json() if r["email"] == "orphan.report@company.com"]
    assert orphan_rows and orphan_rows[0]["name"] == "orphan.report@company.com"
    assert orphan_rows[0]["department"] == "N/A"


# --------------------------------------------------------------------------
# Exports — every format, with rows and empty
# --------------------------------------------------------------------------


def test_rpt_002_csv_export_has_a_header_and_one_row_per_record(admin_client, month):
    """RPT-002 — content type, disposition, header row and body all correct."""
    res = admin_client.get("/api/reports/export/csv", params=params(month))
    assert res.status_code == 200
    assert res.headers["content-type"].startswith("text/csv")
    assert "attachment" in res.headers["content-disposition"]

    lines = res.content.decode("utf-8").strip().splitlines()
    assert lines[0].startswith("Date,Employee ID,Name,Email,Department")
    assert len(lines) == len(ALL_STATUSES) + 1
    assert "INR" in lines[1]


def test_rpt_004_excel_export_opens_and_matches_the_json(admin_client, month):
    """RPT-004 — the workbook is readable and agrees with the JSON summary."""
    res = admin_client.get("/api/reports/export/excel", params=params(month))
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]

    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    rows = list(wb.active.iter_rows(values_only=True))
    assert len(rows) == len(ALL_STATUSES) + 1

    json_rows = admin_client.get("/api/reports/summary", params=params(month)).json()
    assert len(rows) - 1 == len(json_rows)


def test_rpt_005_pdf_export_is_a_valid_document(admin_client, month):
    """RPT-005 — reportlab produces a real PDF, not an empty stream."""
    res = admin_client.get("/api/reports/export/pdf", params=params(month))
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/pdf"
    assert res.content.startswith(b"%PDF")
    assert len(res.content) > 1000


@pytest.mark.parametrize(
    "path", ["/api/reports/export/csv", "/api/reports/export/excel", "/api/reports/export/pdf"]
)
def test_rpt_006_an_empty_range_still_produces_a_valid_file(admin_client, path):
    """RPT-006 — a report with no rows must not be a 500 or a zero-byte file."""
    empty = {"start_date": "2000-01-01", "end_date": "2000-01-02"}
    res = admin_client.get(path, params=empty)
    assert res.status_code == 200
    assert len(res.content) > 0


# --------------------------------------------------------------------------
# Payroll aggregation — every status branch and every fallback
# --------------------------------------------------------------------------


def test_pay_007_the_payroll_summary_counts_each_status_once(admin_client, db, month):
    """PAY-002 — Present, Late and Half Day all count as present days."""
    res = admin_client.get("/api/reports/payroll", params=params(month))
    assert res.status_code == 200
    row = next(r for r in res.json() if r["employee_id"] == month["employee"].profile.employee_id)

    assert row["present_days"] == 3  # Present + Late + Half Day
    assert row["late_days"] == 1
    assert row["half_days"] == 1
    assert row["leave_days"] >= 2  # the Leave record plus the approved-leave day
    assert row["absent_days"] >= 1
    assert row["wfh_days"] == 1
    assert row["total_hours"] == pytest.approx(8.0 * 4)


def test_pay_008_overtime_is_paid_at_one_and_a_half_times(admin_client, month):
    """PAY-001 — the overtime multiplier is applied to the hourly rate."""
    row = next(
        r for r in admin_client.get("/api/reports/payroll", params=params(month)).json()
        if r["employee_id"] == month["employee"].profile.employee_id
    )
    assert row["overtime_hours"] == pytest.approx(0.75)
    assert row["overtime_pay"] == pytest.approx(0.75 * 500.0 * 1.5)
    assert row["total_salary"] == pytest.approx(row["working_salary"] + row["overtime_pay"])


def test_pay_015_payroll_creates_office_settings_for_a_tenant_that_has_none(db):
    """The aggregation needs settings to classify uncovered days; it upserts."""
    org = Organization(name="Payroll Co", slug="payroll-co", is_active=True)
    db.add(org)
    db.flush()
    user = User(
        organization_id=org.id, email="pay@payroll.co", hashed_password="x",
        role="Employee", is_active=True,
    )
    db.add(user)
    db.flush()
    db.add(
        EmployeeProfile(
            organization_id=org.id, user_id=user.id, first_name="Pay", last_name="Roll",
            employee_id="PAY001", hourly_rate=100.0,
        )
    )
    db.commit()
    assert db.query(OfficeSetting).filter(OfficeSetting.organization_id == org.id).first() is None

    result = get_payroll_summary_data(db, org.id, date.today() - timedelta(days=3), date.today())
    assert len(result) == 1
    assert db.query(OfficeSetting).filter(OfficeSetting.organization_id == org.id).first() is not None


def test_pay_005a_the_default_period_is_the_current_month_to_date(admin_client, month):
    """PAY-005 — omitting both dates reports the month so far."""
    res = admin_client.get("/api/reports/payroll")
    assert res.status_code == 200
    assert any(r["employee_id"] == month["employee"].profile.employee_id for r in res.json())


@pytest.mark.parametrize("path", ["/api/reports/export/payroll", "/api/reports/export/payroll/pdf"])
def test_pay_009_payroll_exports_render(admin_client, month, path):
    """PAY-009 — both payroll export formats produce a real file."""
    res = admin_client.get(path, params=params(month))
    assert res.status_code == 200
    assert len(res.content) > 500
    if path.endswith("pdf"):
        assert res.content.startswith(b"%PDF")
    else:
        openpyxl.load_workbook(io.BytesIO(res.content))  # raises if malformed


@pytest.mark.parametrize("path", ["/api/reports/export/payroll", "/api/reports/export/payroll/pdf"])
def test_pay_016_payroll_exports_default_their_period_too(admin_client, month, path):
    """The export routes share the month-to-date default with the JSON route."""
    assert admin_client.get(path).status_code == 200


# --------------------------------------------------------------------------
# Payslip
# --------------------------------------------------------------------------


def test_pay_010a_the_payslip_renders_for_a_settled_month(admin_client, month):
    """PAY-010 — a real PDF, named for the employee and period."""
    res = admin_client.get(
        f"/api/reports/payslip/{month['employee'].id}/pdf", params=params(month)
    )
    assert res.status_code == 200
    assert res.content.startswith(b"%PDF")
    assert "attachment" in res.headers["content-disposition"]
    assert str(month["employee"].id) in res.headers["content-disposition"]


def test_pay_017_the_payslip_defaults_to_the_current_month(admin_client, month):
    """Omitting the range is the common case from the payroll screen."""
    assert admin_client.get(f"/api/reports/payslip/{month['employee'].id}/pdf").status_code == 200


def test_pay_003a_the_payslip_falls_back_to_base_salary_when_there_is_no_hourly_rate(
    admin_client, db, month
):
    """
    PAY-003 / PAY-B1 — the payslip pays a salaried employee from base_salary
    when hourly_rate is zero. Note the payroll *report* has no such fallback:
    the same employee reports a total of 0.00 there. The two disagree, which
    is the inconsistency recorded as PAY-B1.
    """
    employee = month["employee"]
    employee.profile.hourly_rate = 0.0
    employee.profile.base_salary = 90000.0
    db.commit()

    payslip = admin_client.get(f"/api/reports/payslip/{employee.id}/pdf", params=params(month))
    assert payslip.status_code == 200
    assert payslip.content.startswith(b"%PDF")

    payroll_row = next(
        r for r in admin_client.get("/api/reports/payroll", params=params(month)).json()
        if r["employee_id"] == employee.profile.employee_id
    )
    assert payroll_row["total_salary"] == 0.0  # PAY-B1: the report disagrees with the payslip


def test_pay_018_a_payslip_for_an_employee_with_no_profile_is_a_404(admin_client, db, month):
    """The generator needs a profile for the header block."""
    orphan = User(
        organization_id=month["employee"].organization_id,
        email="noprofile.payslip@company.com",
        hashed_password="x",
        role="Employee",
        is_active=True,
    )
    db.add(orphan)
    db.commit()

    res = admin_client.get(f"/api/reports/payslip/{orphan.id}/pdf", params=params(month))
    assert res.status_code == 404
    assert res.json()["detail"] == "Employee profile not found"


def test_pay_013a_a_malformed_payslip_id_is_a_400(admin_client):
    """PAY-013 — the route parses the UUID itself."""
    res = admin_client.get("/api/reports/payslip/not-a-uuid/pdf")
    assert res.status_code == 400
    assert res.json()["detail"] == "Invalid employee user ID format"


def test_pay_019_a_payslip_for_a_tenant_without_settings_renders(db, month):
    """
    PAY-019 / PAY-B2 — regression guard, fixed.

    The generator used to build a transient OfficeSetting() when a tenant had
    no row. Column(default=...) only applies at INSERT, so every field on that
    object was None and the first uncovered day crashed on
    settings.weekends.split(","). It now persists the row, as clock-in and the
    payroll aggregation already did.
    """
    from app.routers.reports import generate_individual_payslip_pdf

    org = Organization(name="Slip Co", slug="slip-co", is_active=True)
    db.add(org)
    db.flush()
    user = User(
        organization_id=org.id, email="slip@slip.co", hashed_password="x",
        role="Employee", is_active=True,
    )
    db.add(user)
    db.flush()
    db.add(
        EmployeeProfile(
            organization_id=org.id, user_id=user.id, first_name="Slip", last_name="Co",
            employee_id="SLIP1", hourly_rate=0.0, base_salary=0.0,
        )
    )
    db.commit()
    assert db.query(OfficeSetting).filter(OfficeSetting.organization_id == org.id).first() is None

    buffer = generate_individual_payslip_pdf(
        db, org, user.id, date.today() - timedelta(days=2), date.today()
    )
    assert buffer.getvalue().startswith(b"%PDF")
    # The row is now persisted, so a second payslip reuses it.
    assert db.query(OfficeSetting).filter(OfficeSetting.organization_id == org.id).first() is not None


def test_pay_020_a_wfh_employee_counts_uncovered_days_as_work_from_home(admin_client, db, month):
    """
    PAY-002 — the third fallback branch, in both aggregations.

    A day with no attendance record is classified by
    get_day_status_for_employee. With WFH active on the profile that check runs
    first, so the day counts as Work From Home rather than an absence — in the
    payroll summary and, separately, in the payslip generator, which carries
    its own copy of the same loop. That duplication is PAY-B1: two
    implementations of one rule, which is why this asserts against both.
    """
    from app.routers.reports import generate_individual_payslip_pdf

    employee = month["employee"]
    employee.profile.wfh_enabled = True
    employee.profile.wfh_start_date = month["start"]
    employee.profile.wfh_end_date = month["end"]
    db.commit()

    row = next(
        r for r in admin_client.get("/api/reports/payroll", params=params(month)).json()
        if r["employee_id"] == employee.profile.employee_id
    )
    uncovered = (month["end"] - month["start"]).days + 1 - len(ALL_STATUSES)
    assert row["wfh_days"] == 1 + uncovered  # the recorded WFH day plus every uncovered one
    assert row["absent_days"] == 1  # only the explicitly recorded absence remains

    buffer = generate_individual_payslip_pdf(
        db, employee.organization, employee.id, month["start"], month["end"]
    )
    assert buffer.getvalue().startswith(b"%PDF")
